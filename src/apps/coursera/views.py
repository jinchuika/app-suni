from django.shortcuts import render
import os
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.urls import reverse_lazy
from django.utils import timezone
from braces.views import (
    LoginRequiredMixin, PermissionRequiredMixin, GroupRequiredMixin,
    CsrfExemptMixin, JsonRequestResponseMixin)
from django.views.generic import DetailView, ListView, View, TemplateView
from django.views.generic.edit import CreateView, UpdateView, FormView
from apps.coursera import models as coursera_m
from apps.coursera import forms as coursera_f
from django.urls import reverse_lazy
from openpyxl import load_workbook
from django.http import HttpResponse, HttpResponseRedirect
from rest_framework import views, status
from rest_framework.response import Response
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Sum

class MonitoreoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Vista para crear una escuela
    """
    model = coursera_m.Monitoreo
    permission_required = "coursera.add_monitoreo"
    template_name = 'coursera/monitoreo_add.html'
    raise_exception = True
    redirect_unauthenticated_users = True
    form_class = coursera_f.FormMonitoreoCrear

    def get_context_data(self, *args, **kwargs):
        context = super(MonitoreoCreateView, self).get_context_data(*args, **kwargs)
        context['registros'] = coursera_m.Monitoreo.objects.all()
        return context

    def get_success_url(self):
        return reverse_lazy('monitoreo_add')

class MonitoreoInformeView(FormView):
    template_name = 'coursera/monitoreo_informe.html'

class CourseraExcelAddView(LoginRequiredMixin, TemplateView):
    template_name = 'coursera/cargar_excel_coursera.html'

    def post(self,request):
        if request.method == 'POST' and request.FILES['myfile']:
            file_list =os.listdir(settings.MEDIA_ROOT_EXCEL_COURSERA)
            myfile=request.FILES['myfile']
            fs=FileSystemStorage(location=settings.MEDIA_ROOT_EXCEL_COURSERA)
            filename=fs.save(myfile.name,myfile)
            uploaded_file_url=fs.url(filename)
            return render(request, 'coursera/cargar_excel_coursera.html',{
                'uploaded_file_url':uploaded_file_url
        })
        return render(request,'bienestar/cargar_excel.html')
class CourseraListView(LoginRequiredMixin, ListView):
    """Vista   para obtener los datos de PrecioEstandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarInformeForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = coursera_m.Monitoreo
    template_name = 'coursera/coursera_informe.html'


class ResultadoCourseraJson(LoginRequiredMixin, views.APIView):
        """ Importar registros de excel con DjangoRestFramework para Bienestar
        """
        def get(self, request):
            file_list =os.listdir(settings.MEDIA_ROOT_EXCEL_COURSERA)
            number_file=len(file_list)
            name_file =file_list[number_file-1]
            #name_file="membership-report.xlsx"
            ruta = str(settings.MEDIA_ROOT_EXCEL_COURSERA)+str(name_file)
            listado_datos=[]
            wb_obj = load_workbook(filename = ruta)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            m_col= sheet_obj.max_column
            for i in range(2, m_row+1):
                try:
                    aliado=sheet_obj.cell(row=i, column=3).value
                except Exception as e:
                    print("esto no funciona")
                if sheet_obj.cell(row=i, column=3).value:
                    external_id=sheet_obj.cell(row=i, column=3).value
                else:
                    external_id="FUNSEPA"
                enrolled_courses=sheet_obj.cell(row=i, column=6).value
                completed_courses=sheet_obj.cell(row=i, column=7).value
                member=sheet_obj.cell(row=i, column=8).value
                if(member=="MEMBER"):
                    nuevo_miembro = 2
                else:
                    nuevo_miembro = 1
                if aliado:
                     if str(aliado).isdigit() or aliado == str("Funsepa"):
                         ingresar_historial=coursera_m.Historial(
                            aliado=coursera_m.Aliado.objects.get(aliado="FUNSEPA"),
                            external_id=external_id,
                            enrolled_courses=enrolled_courses,
                            completed_courses=completed_courses,
                            member=nuevo_miembro
                         )
                         ingresar_historial.save()
                     else:
                         print(i)
                         new_aliado=aliado.split("-")[0]
                         ingresar_historial=coursera_m.Historial(
                           aliado=coursera_m.Aliado.objects.get(aliado=new_aliado),
                           external_id=external_id,
                           enrolled_courses=enrolled_courses,
                           completed_courses=completed_courses,
                           member=nuevo_miembro
                         )
                         ingresar_historial.save()
                else:
                    ingresar_historial=coursera_m.Historial(
                       aliado=coursera_m.Aliado.objects.get(aliado="FUNSEPA"),
                       external_id=external_id,
                       enrolled_courses=enrolled_courses,
                       completed_courses=completed_courses,
                       member=nuevo_miembro
                    )
                    ingresar_historial.save()
            aliado_buscar = coursera_m.Aliado.objects.all()
            for nuevo_aliado in aliado_buscar:
                cuerpo_datos={}
                numero_invitados=coursera_m.Historial.objects.filter(aliado=nuevo_aliado).count()
                numero_miembros=coursera_m.Historial.objects.filter(aliado=nuevo_aliado,member=2).count()
                try:
                	aceptacion= (numero_miembros / numero_invitados)*100
                except ZeroDivisionError:
                    aceptacion=0
                inscritos=coursera_m.Historial.objects.filter(aliado=nuevo_aliado,enrolled_courses__gt=0).count()
                graduados=coursera_m.Historial.objects.filter(aliado=nuevo_aliado,completed_courses__gt=0).count()
                #print("Aliado: "+str(nuevo_aliado)+" Invitaciones: "+str(numero_invitados)+"Miembros: "+str(numero_miembros)+"Porcentaje: "+str(aceptacion)+"Inscritos: "+str(inscritos)+"Graduados: "+str(graduados))
                nuevo_registro = coursera_m.Monitoreo(
                    aliado = nuevo_aliado,
                    invitaciones=numero_invitados,
                    miembros= numero_miembros,
                    aceptacion=aceptacion,
                    inscritos=inscritos,
                    graduados=graduados
                )
                nuevo_registro.save()
            coursera_m.Historial.objects.all().delete()
            return Response(
                "Registros ingresados correctamente",
                status=status.HTTP_200_OK
            )
class ResultadoCourseraMonitoreoJson(LoginRequiredMixin, views.APIView):
        """ Importar registros de excel con DjangoRestFramework para Bienestar
        """
        def get(self, request):
            listado_datos2=[]
            nuevo_monitoreo = coursera_m.Monitoreo.objects.all()
            for datos in nuevo_monitoreo:
                cuerpo_datos2={}
                total_invitaciones =coursera_m.Monitoreo.objects.aggregate(total_invi=Sum('invitaciones'))
                total_miembros =coursera_m.Monitoreo.objects.aggregate(total_miem=Sum('miembros'))
                total_aceptacion =(total_miembros['total_miem']/total_invitaciones['total_invi'])*100
                total_inscrito =coursera_m.Monitoreo.objects.aggregate(total_ins=Sum('inscritos'))
                total_graduados =coursera_m.Monitoreo.objects.aggregate(total_gradu=Sum('graduados'))
                cuerpo_datos2['aliado']=str(datos.aliado)
                cuerpo_datos2['invitaciones']=datos.invitaciones
                cuerpo_datos2['miembros']=datos.miembros
                cuerpo_datos2['aceptacion']=datos.aceptacion
                cuerpo_datos2['inscritos']=datos.inscritos
                cuerpo_datos2['graduados']=datos.graduados
                cuerpo_datos2['total_invitaciones']=total_invitaciones['total_invi']
                cuerpo_datos2['total_miembros']=total_miembros['total_miem']
                cuerpo_datos2['total_aceptacion']=total_aceptacion
                cuerpo_datos2['total_inscrito']=total_inscrito['total_ins']
                cuerpo_datos2['total_graduados']=total_graduados['total_gradu']
                listado_datos2.append(cuerpo_datos2)
            return Response(
                listado_datos2,
                status=status.HTTP_200_OK
            )
