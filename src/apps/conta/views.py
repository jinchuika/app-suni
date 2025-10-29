from django.shortcuts import render
from django.urls import reverse_lazy
from braces.views import (
    LoginRequiredMixin, PermissionRequiredMixin, GroupRequiredMixin,
    CsrfExemptMixin, JsonRequestResponseMixin)
from django.views.generic import DetailView, ListView, View, TemplateView
from django.views.generic.edit import CreateView, UpdateView, FormView
from apps.conta import models as conta_m
from apps.inventario import models as inv_m
from apps.beqt import models as beqt_m
from apps.crm import models as crm_m
from apps.conta import forms as conta_f
from rest_framework import views, status
from rest_framework.response import Response
from django.core.exceptions import ObjectDoesNotExist
from dateutil.relativedelta import relativedelta
from django.utils.datastructures import MultiValueDictKeyError
from datetime import datetime, timedelta
from django.db.models import Sum, Count
from django.db.models import Q
from functools import reduce
from operator import __and__ as AND
from django.db import connection
import re
from apps.escuela import models as escuela_m
from apps.main import creacion_filtros_informe as crear_dict
from django.http import JsonResponse
from django.core import mail
from django.template.loader import render_to_string


import os
import pathlib
import shutil
import platform
import errno
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
import pdfkit
from num2words import num2words
import zipfile
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from openpyxl import load_workbook
from apps.conta.config import EMAIL_CREDENCIALES
import requests
import json
from django.urls import reverse
from apps.main import creacion_filtros_informe as crear_dict

"""
    Función que devuelve la existencia y saldo monetario basado en el tipo de dispositivo,
    fecha y periodo a buscar.
"""
def get_existencia(tipo_dispositivo, fecha, periodo):  
    result = {}
    periodos_anteriores = conta_m.PeriodoFiscal.objects.filter(fecha_fin__lte=periodo.fecha_fin).values('id')

    bajas = conta_m.MovimientoDispositivo.objects.filter(
        tipo_movimiento=-1,
        dispositivo__tipo=tipo_dispositivo,
        fecha__lte=fecha).values('dispositivo')
    compras = conta_m.MovimientoDispositivo.objects.filter(
        tipo_movimiento=1,
        dispositivo__tipo=tipo_dispositivo,
        fecha__lte=fecha,
        dispositivo__entrada__tipo__contable=True).exclude(dispositivo__in=bajas).values('dispositivo')
    utiles = conta_m.MovimientoDispositivo.objects.filter(
        tipo_movimiento=1,
        dispositivo__tipo=tipo_dispositivo,
        fecha__lte=fecha).exclude(
        dispositivo__in=bajas).exclude(dispositivo__in=compras).values('dispositivo')    
        
    # Obtener Precio Estandar Actual y Anterior
       
    if periodo.fecha_fin.year <2023:        
        precio = conta_m.PrecioEstandar.objects.filter(
            tipo_dispositivo=tipo_dispositivo,
            periodo=periodo,
            inventario='dispositivo').first().precio       
    else: 
        precio_lote = conta_m.PrecioDispositivo.objects.filter(
        dispositivo__in=utiles,
        activo=True).aggregate(Sum('precio')) 
        precio = precio_lote['precio__sum']   
        
    # Obtener Precio Total
    # Se valido para los precios de los lotes
    if periodo.fecha_fin.year <= 2018:
        precio_tipo_dispositivo = conta_m.PrecioDispositivo.objects.filter(
            dispositivo__in=utiles,
            periodo__in=periodos_anteriores).aggregate(Sum('precio'))        
    else:
        if periodo.fecha_fin.year <2023:
            #control el precio de los dispositivos de 2023 para bajo       
            precio_tipo_dispositivo = conta_m.PrecioDispositivo.objects.filter(
            dispositivo__in=utiles,
            periodo=periodo).aggregate(Sum('precio'))
        else:
            #controla el precio de los dispositivos arriba del anio 2023
            precio_tipo_dispositivo = conta_m.PrecioDispositivo.objects.filter(
            dispositivo__in=utiles,
            activo=True).aggregate(Sum('precio'))                        
    precio_tipo_compras = conta_m.PrecioDispositivo.objects.filter(
        dispositivo__in=compras,
        activo=True).aggregate(Sum('precio'))

    if precio_tipo_dispositivo['precio__sum'] is not None:       
        precio_tipo_dispositivo = precio_tipo_dispositivo['precio__sum']        
    else:
        if precio is not None:
            if periodo.fecha_fin.year < 2023:
                precio_tipo_dispositivo = len(utiles) * precio              
            else:
                precio_tipo_dispositivo = precio                
        else:
            precio_tipo_dispositivo = 0  
    if precio_tipo_compras['precio__sum'] is not None:
        precio_tipo_compras = precio_tipo_compras['precio__sum']
    else:
        precio_tipo_compras = 0
    precio_total = precio_tipo_dispositivo + precio_tipo_compras    

    existencia = len(utiles) + len(compras)
    result['saldo_total'] = precio_total
    result['existencia'] = existencia
    result['precio_estandar'] = precio
    return result

#Existencias para BEQT

def get_existencia_beqt(tipo_dispositivo, fecha):   
    result = {}

    bajas = conta_m.MovimientoDispositivoBeqt.objects.filter(
        tipo_movimiento=-1,
        dispositivo__tipo=tipo_dispositivo,
        fecha__lte=fecha).values('dispositivo')
    altas = conta_m.MovimientoDispositivoBeqt.objects.filter(
        tipo_movimiento=1,
        dispositivo__tipo=tipo_dispositivo,
        fecha__lte=fecha).exclude(dispositivo__in=bajas).values('dispositivo')
    utiles = conta_m.MovimientoDispositivoBeqt.objects.filter(
        tipo_movimiento=1,
        dispositivo__tipo=tipo_dispositivo,
        fecha__lte=fecha).exclude(
        dispositivo__in=bajas).exclude(dispositivo__in=altas).values('dispositivo')  

    # Obtener Precio Total
    #print("utiles: ",utiles)
    #print("bajas:",bajas)
    #print("altas:", altas)

    #precio_total = precio_tipo_dispositivo + precio_tipo_compras
    existencia = len(utiles) + len(altas)
    precio_total = existencia * 1

    result['saldo_total'] = precio_total
    result['existencia'] = existencia
    #result['precio_estandar'] = precio
    return result

class PeriodoFiscalCreateView(LoginRequiredMixin, CreateView):
    """ Vista   para obtener los datos de Periodo Fiscal mediante una :class:`PeriodoFiscal`
    Funciona  para recibir los datos de un  'PeriodoFiscalForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PeriodoFiscal
    template_name = 'conta/periodo_add.html'
    form_class = conta_f.PeriodoFiscalForm

    def get_success_url(self):
        return reverse_lazy('periodo_list')


class PeriodoFiscalListView(LoginRequiredMixin, ListView):
    """ Vista para Los listados de :class:`PeriodoFiscal`. con sus respectivos datos
    """
    model = conta_m.PeriodoFiscal
    template_name = 'conta/periodo_list.html'


class PeriodoFiscalDetailView(LoginRequiredMixin, DetailView):
    """Vista para detalle de :class:`PeriodoFiscal`.
    """
    model = conta_m.PeriodoFiscal
    template_name = 'conta/periodo_detail.html'


class PeriodoFiscalUpdateView(LoginRequiredMixin, UpdateView):
    """ Vista   para obtener los datos de Periodo Fiscal mediante una :class:`PeriodoFiscal`
    Funciona  para recibir los datos de un  'PeriodoFiscalUpdateForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PeriodoFiscal
    template_name = 'conta/periodo_edit.html'
    form_class = conta_f.PeriodoFiscalUpdateForm

    def get_success_url(self):
        return reverse_lazy('periodo_list')


class PrecioEstandarCreateView(LoginRequiredMixin, CreateView):
    """Vista   para obtener los datos de Precio Estandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/precioestandar_add.html'
    form_class = conta_f.PrecioEstandarForm

    def form_valid(self, form):
        periodo_activo = conta_m.PeriodoFiscal.objects.get(actual=True)
        form.instance.creado_por = self.request.user
        form.instance.periodo = periodo_activo
        return super(PrecioEstandarCreateView, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy('precioestandar_list')


class PrecioEstandarListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de PrecioEstandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarInformeForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/precioestandar_list.html'
    form_class = conta_f.PrecioEstandarInformeForm


class PrecioEstandarInformeListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de PrecioEstandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarInformeForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/informe_existencia.html'
    form_class = conta_f.CantidadInformeForm


class PeriofoFiscalPorExistenciaInformeListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de PrecioEstandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarInformeForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/precioestandar_informe.html'
    form_class = conta_f.PrecioEstandarInformeForm


class ContabilidadEntradaInformeListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de PrecioEstandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarInformeForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/informe_entrada.html'
    form_class = conta_f.EntradaInformeForm

    def get_form(self, form_class=None):
        form = super(ContabilidadEntradaInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = inv_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form

class ContabilidadEntradaDispInformeListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de Dispositivos x Entrada mediante una :class:`PrecioEstandar`
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/informe_entrada_dispositivo.html'
    form_class = conta_f.EntradaDispositivoInformeForm

    def get_form(self, form_class=None):
        form = super(ContabilidadEntradaDispInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = inv_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form

class ContabilidadSalidasInformeListView(LoginRequiredMixin, FormView):
    """Vista utilizada para listar las salidas por rango de fechas y tipo de dispositivo.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/informe_salidas.html'
    form_class = conta_f.SalidasInformeForm
    def get_form(self, form_class=None):
        form = super(ContabilidadSalidasInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = inv_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form

class ContabilidadDesechoInformeListView(LoginRequiredMixin, FormView):
    """Vista utilizada para listar las salidas por rango de fechas y tipo de dispositivo.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/informe_desecho.html'
    form_class = conta_f.DesechoInformeForm

    def get_form(self, form_class=None):
        form = super(ContabilidadDesechoInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = inv_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form

class ContabilidadResumenInformeListView(LoginRequiredMixin, FormView):
    """Vista utilizada para listar el resumen de Inventario.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/informe_resumen.html'
    form_class = conta_f.ResumenInformeForm

    def get_form(self, form_class=None):
        form = super(ContabilidadResumenInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = inv_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form



class ContabilidadResumenPrint(TemplateView):
    template_name = 'conta/informe_resumen_print.html'

    def post(self, request, *args, **kwargs):
        fecha_min = request.POST.get('fecha_min')
        fecha_max = request.POST.get('fecha_max')

        datos_informe_json = request.POST.get('datos_informe_json')
        
        datos_api = []
        error = None
        
        if datos_informe_json:
            try:
                datos_api = json.loads(datos_informe_json)
            except json.JSONDecodeError:
                error = "Error al decodificar los datos del informe enviados."
            except TypeError:
                error = "Los datos del informe enviados no tienen un formato válido."
        else:
            error = "No se recibieron datos para generar el informe." 

        rango_fechas = '{} al {}'.format(fecha_min, fecha_max)

        if datos_api and isinstance(datos_api, list) and datos_api[0]:
             datos_api[0]['rango_fechas'] = rango_fechas
        
        context = {
            'datos_api': datos_api,
            'error': error,
        }
        return render(request, self.template_name, context)

class InformeCantidadJson(views.APIView):

    def get(self, request):
        repuesto_dispositivo = self.request.GET['dispositivo']
        id_periodo = self.request.GET['periodo']
        dispositivos = inv_m.DispositivoTipo.objects.all().exclude(conta=False)
        lista_dispositivos = {}
        lista = []
        acumulador_anterior = 0
        acumulador = 0
        periodo = conta_m.PeriodoFiscal.objects.get(id=id_periodo)
        nueva_fecha = periodo.fecha_fin - timedelta(days=365)
        nueva_fecha_inicio = periodo.fecha_inicio - timedelta(days=365)
        periodo_anterior = conta_m.PeriodoFiscal.objects.get(fecha_fin=nueva_fecha)
        periodos_anteriores = conta_m.PeriodoFiscal.objects.filter(fecha_fin__lt=periodo.fecha_inicio).values('id')
        try:
            for tipo in dispositivos:
                precio_total_anterior = 0
                precio_tipo_dispositivo = 0
                precio_tipo_compras = 0
                dispositivo = {}
                if repuesto_dispositivo == str(1):
                    # Obtener Total Anterior
                    totales_anterior = get_existencia(tipo,nueva_fecha,periodo_anterior)
                    precio_anterior = totales_anterior['precio_estandar']
                    precio_total_anterior = totales_anterior['saldo_total']
                    acumulador_anterior += precio_total_anterior

                    # Obtener Precio Estandar Actual y Anterior
                    total_actual = get_existencia(tipo,nueva_fecha,periodo)
                    precio = total_actual['precio_estandar']
                    precio_total = total_actual['saldo_total']
                    acumulador += precio_total

                    # Obtener Existencia
                    disponible = total_actual['existencia']
                else:
                    bajas = conta_m.MovimientoRepuesto.objects.filter(
                        tipo_movimiento=-1,
                        repuesto__tipo=tipo,
                        fecha__lte=nueva_fecha).values('repuesto')
                    compras = conta_m.MovimientoRepuesto.objects.filter(
                        tipo_movimiento=1,
                        repuesto__tipo=tipo,
                        fecha__lte=nueva_fecha,
                        repuesto__entrada__tipo__contable=True).exclude(repuesto__in=bajas).values('repuesto')
                    utiles = conta_m.MovimientoRepuesto.objects.filter(
                        tipo_movimiento=1,
                        repuesto__tipo=tipo,
                        fecha__lte=nueva_fecha).exclude(
                            repuesto__in=bajas).exclude(repuesto__in=compras).values('repuesto')
                    # Obtener Total Anterior
                    if periodo_anterior.fecha_fin.year <= 2018:
                        precio_tipo_dispositivo = conta_m.PrecioRepuesto.objects.filter(
                            repuesto__in=utiles,
                            periodo__in=periodos_anteriores).aggregate(Sum('precio'))
                    else:
                        precio_tipo_dispositivo = conta_m.PrecioRepuesto.objects.filter(
                            repuesto__in=utiles,
                            periodo=periodo_anterior).aggregate(Sum('precio'))

                    precio_tipo_compras = conta_m.PrecioRepuesto.objects.filter(
                        repuesto__in=compras,
                        activo=True).aggregate(Sum('precio'))

                    if precio_tipo_dispositivo['precio__sum'] is not None:
                        precio_tipo_dispositivo = precio_tipo_dispositivo['precio__sum']
                    else:
                        precio_tipo_dispositivo = 0

                    if precio_tipo_compras['precio__sum'] is not None:
                        precio_tipo_compras = precio_tipo_compras['precio__sum']
                    else:
                        precio_tipo_compras = 0
                    precio_total_anterior = precio_tipo_dispositivo + precio_tipo_compras
                    acumulador_anterior += precio_total_anterior
                    # Obtener Precio Estandar Actual y Anterior
                    precio = conta_m.PrecioEstandar.objects.filter(
                        tipo_dispositivo=tipo,
                        periodo=periodo,
                        inventario='repuesto').first().precio

                    precio_anterior = conta_m.PrecioEstandar.objects.filter(
                        tipo_dispositivo=tipo,
                        periodo=periodo_anterior,
                        inventario='repuesto').first().precio
                    # Obtener Total Actual
                    precio_tipo_dispositivo = conta_m.PrecioRepuesto.objects.filter(
                        repuesto__in=utiles,
                        periodo=periodo).aggregate(Sum('precio'))
                    precio_tipo_compras = conta_m.PrecioRepuesto.objects.filter(
                        repuesto__in=compras,
                        activo=True).aggregate(Sum('precio'))

                    if precio_tipo_dispositivo['precio__sum'] is not None:
                        precio_tipo_dispositivo = precio_tipo_dispositivo['precio__sum']
                    else:
                        if precio is not None:
                            precio_tipo_dispositivo = len(utiles) * precio
                        else:
                            precio_tipo_dispositivo = 0

                    if precio_tipo_compras['precio__sum'] is not None:
                        precio_tipo_compras = precio_tipo_compras['precio__sum']
                    else:
                        precio_tipo_compras = 0

                    precio_total = precio_tipo_dispositivo + precio_tipo_compras
                    acumulador += precio_total
                    disponible = len(utiles) + len(compras)
                if precio is not None and precio_anterior is not None:
                    dispositivo['tipo'] = tipo.tipo
                    dispositivo['cantidad'] = disponible
                    dispositivo['precio'] = str(precio)
                    dispositivo['precio_anterior'] = precio_anterior
                    dispositivo['total_anterior'] = precio_total_anterior
                    dispositivo['total'] = precio_total
                    dispositivo['acumulador_total'] = acumulador
                    dispositivo['acumulador_anterior'] = acumulador_anterior
                    lista.append(dispositivo)
                    lista_dispositivos[tipo.tipo] = dispositivo
            return Response(lista)
        except ObjectDoesNotExist as e:
            return Response(
                {
                    'mensaje': str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )


class InformeEntradaJson(views.APIView):
    """ Si es Donacion o Contenedor ir a traer el precio estandar del modelo de contabilidad
    en caso contrario desde el precio_unitario del detalle de entrada
    """
    def get(self, request):
        try:
            donante = self.request.GET['donante']
        except MultiValueDictKeyError as e:
            donante = 0
        try:
            tipo_entrada = []
            tipo_entrada = request.GET.getlist('tipo_entrada[]')
            if len(tipo_entrada) == 0:
                tipo = self.request.GET['tipo_entrada']
                tipo_entrada.append(tipo)
        except MultiValueDictKeyError as e:
            tipo_entrada = 0
        tipo_dispositivo = self.request.GET['tipo_dispositivo']
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        tipo_dispositivo_nombre = inv_m.DispositivoTipo.objects.get(pk=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        validar_fecha = conta_m.PeriodoFiscal.objects.filter(fecha_inicio__lte=fecha_inicio, fecha_fin__gte=fecha_fin)
        if validar_fecha.count() == 1:

            # Obtener datos de Periodo Fiscal
            periodo = validar_fecha[0]
            lista_dispositivos = {}
            lista = []

            # Preparar Filtros de Búsqueda
            q = []
            q.append(Q(tipo_dispositivo=tipo_dispositivo))

            if donante and donante != 0:
                q.append(Q(entrada__proveedor=donante))

            if tipo_entrada and tipo_entrada != 0:
                q.append(Q(entrada__tipo__in=tipo_entrada))

            # obtener Listado de Detalles de Entrada Aplicando Filtros
            entrada_detalle = inv_m.EntradaDetalle.objects.values('entrada','precio_unitario').filter(
                        Q(fecha_dispositivo__gte=fecha_inicio),
                        Q(fecha_dispositivo__lte=fecha_fin),
                        reduce(AND,q)
                        ).exclude(entrada__tipo__nombre='Especial').annotate(Sum('util'),Sum('precio_unitario'))
            
            # Obtener Existencia Inicial y Saldo Inicial
            fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
            totales_anterior = get_existencia(tipo_dispositivo,fecha_inicial,periodo)
            precio_anterior = totales_anterior['precio_estandar']
            precio_total_anterior = totales_anterior['saldo_total']
            existencia_anterior = totales_anterior['existencia']

            # Obtener Existencia Final y Saldo Final
            total_actual = get_existencia(tipo_dispositivo,fecha_fin,periodo)
            precio_actual = total_actual['precio_estandar']
            precio_total = total_actual['saldo_total']
            existencia_actual = total_actual['existencia']
            for datos_entrada in entrada_detalle:
                entrada = inv_m.Entrada.objects.get(pk=datos_entrada['entrada'])
                #precio = datos_entrada['precio_unitario__sum']
                precio = datos_entrada['precio_unitario']
                cantidad = datos_entrada['util__sum']
                dispositivo = {}

                # Validar Precio de Compra y Donación
                #print(entrada.tipo)
                if (not precio or precio == 0) and not entrada.tipo.contable:
                    # Obtener Precio Estandar
                    #print("Ingreso aca 1")
                    precio = conta_m.PrecioEstandar.objects.get(
                    tipo_dispositivo=tipo_dispositivo,
                    periodo=periodo,
                    inventario="dispositivo").precio
                    
                  
                else:
                    #precio = 0                    
                    print("Ingreso aca 22")
                dispositivo['fecha'] = entrada.fecha
                dispositivo['id'] = entrada.id
                dispositivo['url'] = entrada.get_absolute_url()
                dispositivo['util'] = cantidad
                dispositivo['precio'] = round(precio,2)
                dispositivo['tipo'] = entrada.tipo.nombre
                dispositivo['proveedor'] = entrada.proveedor.nombre
                try:
                    dispositivo['total'] = round(cantidad * precio,2)                    
                except:
                    print(entrada.id)
                dispositivo['total_costo'] = precio_total_anterior
                dispositivo['total_final'] = existencia_anterior
                dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
                dispositivo['tipo_dispositivo'] = tipo_dispositivo_nombre.tipo
                dispositivo['total_costo_despues'] = precio_total
                dispositivo['total_despues'] = existencia_actual
                lista.append(dispositivo)
            return Response(lista)
        else:
            print("No existe en el periodo fiscal")

class InformeEntradaDispositivoJson(views.APIView):
    """ Lista los dispositivos ingresados al sistema por rango de fechas
    """
    def get(self, request):
        try:
            no_entrada = self.request.GET['no_entrada']
        except MultiValueDictKeyError as e:
            no_entrada = 0

        tipo_dispositivo = self.request.GET['tipo_dispositivo']
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        tipo_dispositivo_nombre = inv_m.DispositivoTipo.objects.get(pk=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        validar_fecha = conta_m.PeriodoFiscal.objects.filter(fecha_inicio__lte=fecha_inicio, fecha_fin__gte=fecha_fin)
        print(validar_fecha)
        if validar_fecha.count() == 1:

            # Obtener datos de Periodo Fiscal
            periodo = validar_fecha[0]
            lista_dispositivos = {}
            lista = []

            # Preparar Filtros de Búsqueda
            q = []
            q.append(Q(tipo=tipo_dispositivo))

            if no_entrada and no_entrada != 0:
                q.append(Q(entrada_detalle__entrada=no_entrada))

            # obtener Listado de Detalles de Entrada Aplicando Filtros
            entrada_detalle = inv_m.Dispositivo.objects.filter(
                        Q(entrada_detalle__fecha_dispositivo__gte=fecha_inicio),
                        Q(entrada_detalle__fecha_dispositivo__lte=fecha_fin),
                        reduce(AND,q)
                        ).exclude(entrada_detalle__entrada__tipo__nombre='Especial')

            print(entrada_detalle)

            # Obtener Existencia Inicial y Saldo Inicial
            fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
            totales_anterior = get_existencia(tipo_dispositivo,fecha_inicial,periodo)
            existencia_anterior = totales_anterior['existencia']

            # Obtener Existencia Final y Saldo Final
            total_actual = get_existencia(tipo_dispositivo,fecha_fin,periodo)
            existencia_actual = total_actual['existencia']

            for detalle in entrada_detalle:
                entrada = inv_m.Entrada.objects.get(pk=detalle.entrada_detalle.entrada.id)
                dispositivo = {}

                dispositivo['triage'] = detalle.triage
                dispositivo['entrada'] = detalle.entrada_detalle.entrada.id
                dispositivo['fecha'] = detalle.entrada_detalle.fecha_dispositivo
                dispositivo['tipo_entrada'] = detalle.entrada_detalle.entrada.tipo.nombre
                dispositivo['url'] = detalle.get_absolute_url()
                dispositivo['url_entrada'] = detalle.entrada_detalle.entrada.get_absolute_url()

                dispositivo['total_final'] = existencia_anterior
                dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
                dispositivo['tipo_dispositivo'] = tipo_dispositivo_nombre.tipo
                dispositivo['total_despues'] = existencia_actual
                lista.append(dispositivo)
            return Response(lista)
        else:
            print("No existe en el periodo fiscal")

class InformeSalidaJson(views.APIView):
    """ Listar las salidas por el tipo seleccionado, si es compra o no traer o excluir los dispositivos de compra del listado.
    """
    def get(self, request):
        try:
            udi = self.request.GET['udi']
        except MultiValueDictKeyError as e:
            udi = 0

        try:
            beneficiado = self.request.GET['beneficiado']
        except MultiValueDictKeyError as e:
            beneficiado = 0

        try:
            donaciones = self.request.GET.get('donaciones','') == 'on'
        except MultiValueDictKeyError as e:
            donaciones = False

        try:
            compra = self.request.GET.get('compras','') == 'on'
        except MultiValueDictKeyError as e:
            compra = False

        try:
            tipo_salida = []
            tipo_salida = request.GET.getlist('tipo_salida[]')
            if len(tipo_salida) == 0:
                tipo = self.request.GET['tipo_salida']
                tipo_salida.append(tipo)
        except MultiValueDictKeyError as e:
            tipo_salida = 0

        tipo_dispositivo = self.request.GET['tipo_dispositivo']
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        tipo_dispositivo_nombre = inv_m.DispositivoTipo.objects.get(pk=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        validar_fecha = conta_m.PeriodoFiscal.objects.filter(fecha_inicio__lte=fecha_inicio, fecha_fin__gte=fecha_fin)
        if validar_fecha.count() == 1:

            # Obtener datos de Periodo Fiscal
            periodo = validar_fecha[0]
            salida_especial = inv_m.SalidaTipo.objects.get(especial=True)
            #print(inv_m.EntradaTipo.objects.filter(contable=True).last())
            tipo_compra =  inv_m.EntradaTipo.objects.filter(contable=True).first()
            lista_dispositivos = {}
            lista = []

            # Armar Query de Consulta
            sql_select = '''SELECT isi.id as 'Salida', COUNT(*) AS 'CANTIDAD', cpd.precio AS 'PRECIO', SUM(cpd.precio) AS 'TOTAL'
                    FROM inventario_dispositivopaquete idp
                    INNER JOIN inventario_dispositivo id on idp.dispositivo_id = id.id
                    INNER JOIN inventario_paquete ip on idp.paquete_id = ip.id
                    INNER JOIN inventario_salidainventario isi on ip.salida_id = isi.id
                    INNER JOIN inventario_entrada ie on id.entrada_id = ie.id
                    INNER JOIN conta_preciodispositivo cpd on idp.dispositivo_id = cpd.dispositivo_id'''

            sql_where = """ WHERE id.tipo_id = {tipo_dispositivo}
                    AND isi.en_creacion = 0
                    AND isi.fecha between '{fecha_inicio}' AND '{fecha_fin}'
                    AND isi.tipo_salida_id <> {especial}
                    AND cpd.activo = 1""".format(tipo_dispositivo = tipo_dispositivo,
                                                                                        fecha_inicio=fecha_inicio,
                                                                                        fecha_fin=fecha_fin,
                                                                                        especial=salida_especial.id)
            sql_group = " GROUP BY isi.id, cpd.precio"

            if udi and udi != 0:
                escuela = escuela_m.Escuela.objects.get(codigo=udi)
                if escuela:
                    sql_where += " AND isi.escuela_id = " + str(escuela.id)

            if beneficiado and beneficiado != 0:
                sql_where += " AND isi.beneficiario_id = " + str(beneficiado)

            if tipo_salida and tipo_salida != 0:
                sql_where += " AND isi.tipo_salida_id in (" + ','.join(tipo_salida) + ")"

            if compra and not donaciones:
                sql_where += " AND ie.tipo_id = " + str(tipo_compra.id)
            elif not compra and donaciones:
                sql_where += " AND ie.tipo_id <> " + str(tipo_compra.id)

            sql_query = sql_select + sql_where + sql_group

            with connection.cursor() as cursor:
                cursor.execute(sql_query)
                result =cursor.fetchall()

            # Obtener Existencia Inicial y Saldo Inicial
            fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
            totales_anterior = get_existencia(tipo_dispositivo,fecha_inicial,periodo)
            precio_anterior = totales_anterior['precio_estandar']
            precio_total_anterior = totales_anterior['saldo_total']
            existencia_anterior = totales_anterior['existencia']

            # Obtener Existencia Final y Saldo Final
            total_actual = get_existencia(tipo_dispositivo,fecha_fin,periodo)
            precio_actual = total_actual['precio_estandar']
            precio_total = total_actual['saldo_total']
            existencia_actual = total_actual['existencia']
            for datos_salida in result:
                salida = inv_m.SalidaInventario.objects.get(pk=datos_salida[0])
                cantidad = datos_salida[1]
                precio = datos_salida[2]
                total = datos_salida[3]
                dispositivo = {}

                if salida.escuela:
                    beneficiado = salida.escuela.nombre
                elif salida.beneficiario:
                    beneficiado = salida.beneficiario.nombre
                else:
                    beneficiado = ''

                dispositivo['fecha'] = salida.fecha
                dispositivo['id'] = salida.id
                dispositivo['no_salida'] = salida.no_salida
                dispositivo['url'] = salida.get_absolute_url()
                dispositivo['util'] = cantidad
                dispositivo['precio'] = precio
                dispositivo['tipo'] = salida.tipo_salida.nombre
                dispositivo['beneficiado'] = beneficiado
                dispositivo['total'] = total
                dispositivo['total_costo'] = precio_total_anterior
                dispositivo['total_final'] = existencia_anterior
                dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
                dispositivo['tipo_dispositivo'] = tipo_dispositivo_nombre.tipo
                dispositivo['total_costo_despues'] = precio_total
                dispositivo['total_despues'] = existencia_actual
                lista.append(dispositivo)
            return Response(lista)
        else:
            print("No existe en el periodo fiscal")

class InformeDesechoJson(views.APIView):
    """ Lista todas las salidas de desecho con triage que han sucedido en un rango de fechas,
    Solamente cuentan aquellas salidas que han sido cerradas.
    """
    def get(self, request):
        try:
            empresa = self.request.GET['empresa']
        except MultiValueDictKeyError as e:
            empresa = 0

        tipo_dispositivo = self.request.GET['tipo_dispositivo']
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        tipo_dispositivo_nombre = inv_m.DispositivoTipo.objects.get(pk=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        validar_fecha = conta_m.PeriodoFiscal.objects.filter(fecha_inicio__lte=fecha_inicio, fecha_fin__gte=fecha_fin)
        if validar_fecha.count() == 1:

            # Obtener datos de Periodo Fiscal
            periodo = validar_fecha[0]
            lista_dispositivos = {}
            lista = []

            # Preparar Filtros de Búsqueda
            q = []
            q.append(Q(dispositivo__tipo=tipo_dispositivo))

            if empresa and empresa != 0:
                q.append(Q(desecho__empresa=empresa))

            # obtener Listado de Detalles de Desecho Aplicando Filtros
            salida_detalle = inv_m.DesechoDispositivo.objects.values('desecho').filter(
                        Q(desecho__fecha__gte=fecha_inicio),
                        Q(desecho__fecha__lte=fecha_fin),
                        Q(aprobado=True),
                        Q(desecho__en_creacion=False),
                        reduce(AND,q)
                        ).annotate(Count('dispositivo'))           

            # Obtener Existencia Inicial y Saldo Inicial
            fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
            totales_anterior = get_existencia(tipo_dispositivo,fecha_inicial,periodo)
            precio_anterior = totales_anterior['precio_estandar']
            precio_total_anterior = totales_anterior['saldo_total']
            existencia_anterior = totales_anterior['existencia']

            # Obtener Existencia Final y Saldo Final
            total_actual = get_existencia(tipo_dispositivo,fecha_fin,periodo)
            precio_actual = total_actual['precio_estandar']
            precio_total = total_actual['saldo_total']
            existencia_actual = total_actual['existencia']

            for datos_desecho in salida_detalle:                
                salida = inv_m.DesechoSalida.objects.get(pk=datos_desecho['desecho'])
                cantidad = datos_desecho['dispositivo__count']
                dispositivo = {}
                salida_dispositivo = inv_m.DesechoDispositivo.objects.filter(desecho=salida,dispositivo__tipo=tipo_dispositivo_nombre)
                print(salida_dispositivo)

                # Obtener Precio Estandar
                precio = conta_m.PrecioEstandar.objects.get(
                    tipo_dispositivo=tipo_dispositivo,
                    periodo=periodo,
                    inventario="dispositivo").precio

                dispositivo['fecha'] = salida.fecha
                dispositivo['id'] = salida.id
                dispositivo['url'] = salida.get_absolute_url()
                dispositivo['util'] = cantidad
                dispositivo['precio'] = precio
                dispositivo['recolectora'] = salida.empresa.nombre
                dispositivo['total'] = cantidad * precio
                dispositivo['total_costo'] = precio_total_anterior
                dispositivo['total_final'] = existencia_anterior
                dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
                dispositivo['tipo_dispositivo'] = tipo_dispositivo_nombre.tipo
                dispositivo['total_costo_despues'] = precio_total
                dispositivo['total_despues'] = existencia_actual  
                lista.append(dispositivo)
            return Response(lista)
        else:
            print("No existe en el periodo fiscal")

class InformeResumenJson(views.APIView):
    """ Lista todas las salidas de desecho con triage que han sucedido en un rango de fechas,
    Solamente cuentan aquellas salidas que han sido cerradas.
    """
    def get(self, request):
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        try:
            tipo_dispositivo = []
            tipo_dispositivo = self.request.GET.getlist('tipo_dispositivo[]')
            if len(tipo_dispositivo) == 0:
                tipo = self.request.GET['tipo_dispositivo']
                tipo_dispositivo.append(tipo)
        except MultiValueDictKeyError as e:
            tipo_dispositivo = 0

        # Filtrar por tipos de dispositivos seleccionados
        if tipo_dispositivo == 0 or not tipo_dispositivo:
            dispositivos = self.request.user.tipos_dispositivos.tipos.filter(conta=True)
        else:
            dispositivos = self.request.user.tipos_dispositivos.tipos.filter(id__in=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        validar_fecha = conta_m.PeriodoFiscal.objects.filter(fecha_inicio__lte=fecha_inicio, fecha_fin__gte=fecha_fin)
        acumulador = 0
        acumulador_anterior = 0
        acumulador_ant_ex = 0
        acumulador_act_ex = 0

        if validar_fecha.count() == 1:
            # Obtener datos de Periodo Fiscal            
            periodo = validar_fecha[0]
            lista_dispositivos = {}
            lista = []
            for tipo in dispositivos:             
                dispositivo = {}
                # Obtener Saldo Anterior
                fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
                totales_anterior = get_existencia(tipo,fecha_inicial,periodo)                
                precio_total_anterior = totales_anterior['saldo_total']
                existencia_anterior = totales_anterior['existencia']
                acumulador_anterior += precio_total_anterior
                acumulador_ant_ex += existencia_anterior

                # Obtener Saldo Actual
                total_actual = get_existencia(tipo,fecha_fin,periodo)                            
                precio = total_actual['precio_estandar']
                precio_total = total_actual['saldo_total']
                existencia = total_actual['existencia']
                acumulador += precio_total
                acumulador_act_ex += existencia
                #print("Tipo >{} , Precio > {}",tipo,precio)
                # Obtener Total de Entradas
                entradas = inv_m.EntradaDetalle.objects.filter(
                        Q(fecha_dispositivo__gte=fecha_inicio),
                        Q(fecha_dispositivo__lte=fecha_fin),
                        Q(tipo_dispositivo=tipo)
                        ).aggregate(Sum('util'))               
                if entradas['util__sum'] is not None:
                    entradas = entradas['util__sum']
                else:
                    entradas = 0

                # Obtener Total de Salidas
                salida_especial = inv_m.SalidaTipo.objects.get(especial=True)
                salidas = len(inv_m.DispositivoPaquete.objects.filter(
                        Q(paquete__salida__fecha__gte=fecha_inicio),
                        Q(paquete__salida__fecha__lte=fecha_fin),
                        Q(dispositivo__tipo=tipo),
                        Q(paquete__salida__en_creacion=False)))

                desecho = len(inv_m.DesechoDispositivo.objects.filter(
                        Q(desecho__fecha__gte=fecha_inicio),
                        Q(desecho__fecha__lte=fecha_fin),
                        Q(dispositivo__tipo=tipo),
                        Q(desecho__en_creacion=False)))

                salidas += desecho

                dispositivo['tipo'] = tipo.tipo
                dispositivo['existencia_anterior'] = existencia_anterior
                dispositivo['saldo_anterior'] = precio_total_anterior
                dispositivo['entradas'] = str(entradas)
                dispositivo['salidas'] = salidas
                dispositivo['existencia'] = existencia
                dispositivo['saldo_actual'] = precio_total
                dispositivo['costo_inicial'] = acumulador_anterior
                dispositivo['total_inicial'] = acumulador_ant_ex
                dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
                dispositivo['costo_final'] = acumulador
                dispositivo['total_final'] = acumulador_act_ex
                lista.append(dispositivo)
            return Response(lista)
        else:
            print("No existe en el periodo fiscal")


class InformeExistencias(views.APIView):
    """ Lista todas los dispositivios que estan existencia por media de la `class:``Movimiento dispositivo`.
    """
    def get(self, request):      
        try:
            fecha_inicio = self.request.GET['fecha_min']
        except MultiValueDictKeyError:
            fecha_inicio=0
        
        try:
            fecha_fin = self.request.GET['fecha_max']
        except MultiValueDictKeyError:
            fecha_fin=0

        try:
            tipo_dispositivo = [int(x) for x in self.request.GET.getlist('tipo_dispositivo[]')]
            if len(tipo_dispositivo) ==0:
                tipo_dispositivo = self.request.GET['tipo_dispositivo']            
        except MultiValueDictKeyError:
            tipo_dispositivo=0
        fecha_vacia = str(fecha_inicio) + str(fecha_fin)        
        if fecha_inicio == 0 :
            rango_fecha = "AL "+str(fecha_fin)
        elif fecha_fin == 0:
            rango_fecha = "DEL "+str(fecha_inicio)        
        else:
            rango_fecha = "Del "+str(fecha_inicio) + "AL "+str(fecha_fin)        
        if  fecha_vacia == 0:
            rango_fecha = "AL "+str(datetime.now().date())  
        saldo_total = 0
        datos_existencia = []
        sort_params ={}
        sort_params_bajas = {}
        crear_dict.crear_dict(sort_params,'dispositivo__tipo_id__in',tipo_dispositivo)
        crear_dict.crear_dict(sort_params,'fecha__gte',fecha_inicio)
        crear_dict.crear_dict(sort_params,'fecha__lte',fecha_fin)
        crear_dict.crear_dict(sort_params,'tipo_movimiento',1)
        #crear_dict.crear_dict(sort_params,'dispositivo__valido',True)
        crear_dict.crear_dict(sort_params,'dispositivo__valido',False)
        crear_dict.crear_dict(sort_params_bajas,'dispositivo__tipo_id__in',tipo_dispositivo)
        crear_dict.crear_dict(sort_params_bajas,'fecha__gte',fecha_inicio)
        crear_dict.crear_dict(sort_params_bajas,'fecha__lte',fecha_fin)
        crear_dict.crear_dict(sort_params_bajas,'tipo_movimiento',-1)
        crear_dict.crear_dict(sort_params_bajas,'dispositivo__valido',False)
        dispositivos_baja  = conta_m.MovimientoDispositivo.objects.filter(**sort_params_bajas)
        dispositivos_alta  = conta_m.MovimientoDispositivo.objects.filter(**sort_params)         
        for data in dispositivos_alta:
            if dispositivos_baja.filter(dispositivo=data.dispositivo).count() ==0:                
                datos_existencia.append(data)            
        dispositivos =[]
        existencias_total = 0
        
        for data in datos_existencia:
            precio_dipositivo = conta_m.PrecioDispositivo.objects.get(dispositivo=data.dispositivo,activo=True)           
            existencias = {}
            existencias["triage"] =data.dispositivo.triage
            existencias["tipo"] = data.dispositivo.tipo.tipo
            existencias["fecha_ingreso"] = data.dispositivo.entrada.fecha
            existencias["tipo_ingreso"] = data.dispositivo.entrada.tipo.nombre
            #existencias["precio"] = data.precio
            existencias["precio"] = precio_dipositivo.precio
            existencias["fecha_precio"] = precio_dipositivo.fecha_creacion
            saldo_total = saldo_total + precio_dipositivo.precio
            existencias["rango_fecha"] = rango_fecha
            existencias["fecha"]= rango_fecha
            existencias_total = existencias_total + 1
            existencias["existencias_total"]= existencias_total
            existencias["saldo_total"]= saldo_total
            dispositivos.append(existencias)  
        return Response(dispositivos)
    
class ExistenciaDispositivosInformeView(LoginRequiredMixin, GroupRequiredMixin, FormView):
    """ Vista para obtener la informacion de los dispositivos para crear el informe de existencia mediante un
    api mediante el metodo GET  y lo muestra en el tempalte
    """
    group_required = [u"inv_conta", ]
    redirect_unauthenticated_users = True
    raise_exception = True
    template_name = "conta/informe_existencia_dispositivos.html"
    form_class = conta_f.ExistenciaDispositivosInformeForm


class InformeRastreoDesecho(views.APIView):
    """ Lista todas los dispositivios que estan en desecho por medio de la `class:`Desecho` de inventario.
    """
    def get(self, request):      
        try:
            fecha_inicio = self.request.GET['fecha_min']
        except MultiValueDictKeyError:
            fecha_inicio=0
        
        try:
            fecha_fin = self.request.GET['fecha_max']
        except MultiValueDictKeyError:
            fecha_fin=0
        
        try:
            fecha_inicio_entrada = self.request.GET['fecha_min_entrada']
        except MultiValueDictKeyError:
            fecha_inicio_entrada=0
        
        try:
            fecha_fin_entrada = self.request.GET['fecha_max_entrada']
        except MultiValueDictKeyError:
            fecha_fin_entrada=0

        try:
            tipo_dispositivo = [int(x) for x in self.request.GET.getlist('tipo_dispositivo[]')]
            if len(tipo_dispositivo) ==0:
                tipo_dispositivo = self.request.GET['tipo_dispositivo']            
        except MultiValueDictKeyError:
            tipo_dispositivo=0

        try:
            tipo_entrada = [int(x) for x in self.request.GET.getlist('tipo_entrada[]')]
            if len(tipo_entrada) ==0:
                tipo_entrada = self.request.GET['tipo_entrada']            
        except MultiValueDictKeyError:
            tipo_entrada=0
        try:
            entradas = [int(x) for x in self.request.GET.getlist('entrada_inventario[]')]
            list_entrada = [] 
            if len(entradas) ==0:
                entradas = self.request.GET['entrada_inventario'] 
                list_entrada.append(entradas)                          
        except MultiValueDictKeyError:
            entradas=0

        try:
            desecho = [int(x) for x in self.request.GET.getlist('entrada[]')]
            if len(desecho) ==0:
                desecho = self.request.GET['entrada']            
        except MultiValueDictKeyError:
            desecho=0   
        sort_params ={}        
        crear_dict.crear_dict(sort_params,'tipo_dispositivo__id__in',tipo_dispositivo)
        crear_dict.crear_dict(sort_params,'desecho__fecha__gte',fecha_inicio)
        crear_dict.crear_dict(sort_params,'desecho__fecha__lte',fecha_fin)
        crear_dict.crear_dict(sort_params,'entrada_detalle__entrada__tipo__id__in',tipo_entrada)
        if len(list_entrada)==1:
            crear_dict.crear_dict(sort_params,'entrada_detalle__entrada__id__in',list_entrada)
        else:
            crear_dict.crear_dict(sort_params,'entrada_detalle__entrada__id__in',entradas)
        crear_dict.crear_dict(sort_params,'entrada_detalle__entrada__fecha__gte',fecha_inicio_entrada)
        crear_dict.crear_dict(sort_params,'entrada_detalle__entrada__fecha__lte',fecha_fin_entrada)
        crear_dict.crear_dict(sort_params,'id__in',desecho)    
        desecho_detalle = inv_m.DesechoDetalle.objects.filter(**sort_params)        
        datos_desecho =[]
        cantidad_total = 0   
        for data in desecho_detalle:
            desecho = {}
            desecho["desecho_id"] =data.desecho.id
            desecho["desecho_url"] =data.desecho.get_absolute_url_detail()
            desecho["desecho_fecha"] = str(data.desecho.fecha)
            desecho["desecho_precio_total"] = data.desecho.precio_total
            desecho["desecho_peso"] = data.desecho.peso
            desecho["desecho_tecnico"] = data.desecho.creado_por.get_full_name()
            desecho["desecho_empresa"] = data.desecho.empresa.nombre           
            desecho["desecho_cantidad"] = data.cantidad
            desecho["desecho_tipo_dispositivo"]= data.tipo_dispositivo.tipo
            desecho["desecho_observaciones"]= data.desecho.observaciones
            desecho["entrada_detalle"]= data.entrada_detalle.id
            desecho["entrada_detalle_util"]= data.entrada_detalle.util
            desecho["entrada_detalle_repuesto"]= data.entrada_detalle.repuesto
            desecho["entrada_detalle_desecho"]= data.entrada_detalle.desecho
            desecho["entrada_detalle_precio_unitario"]= data.entrada_detalle.precio_unitario
            desecho["entrada_detalle_precio_total"]= data.entrada_detalle.precio_total
            desecho["entrada_detalle_tecnico"]= data.entrada_detalle.creado_por.get_full_name()
            desecho["entrada_detalle_entrada_id"]= data.entrada_detalle.entrada.id
            desecho["entrada_detalle_descripcion"]= data.entrada_detalle.descripcion
            desecho["entrada_detalle_fecha"]= str(data.entrada_detalle.entrada.fecha)
            desecho["entrada_detalle_total"]= data.entrada_detalle.total            
            desecho["entrada_proveedor"]= data.entrada_detalle.entrada.proveedor.nombre
            desecho["entrada_tecnico"]= data.entrada_detalle.entrada.creada_por.get_full_name()
            desecho["entrada_factura"]= data.entrada_detalle.entrada.factura
            desecho["entrada_fecha_cierre"]= str(data.entrada_detalle.entrada.fecha_cierre)
            desecho["entrada_tipo"]= data.entrada_detalle.entrada.tipo.nombre 
            desecho["entrada_url"] =data.entrada_detalle.entrada.get_absolute_url()
            desecho["proveedor_url"] =data.desecho.empresa.get_absolute_url_detail()
            cantidad_total = cantidad_total +  data.cantidad
            desecho["cantidad_total"] =cantidad_total
            datos_desecho.append(desecho)   
        return Response(datos_desecho)
    
class DesechoRastreoInformeView(LoginRequiredMixin, GroupRequiredMixin, FormView):
    """ Vista para obtener la informacion de los dispositivos para crear el informe de existencia mediante un
    api mediante el metodo GET  y lo muestra en el tempalte
    """
    group_required = [u"inv_conta",u"inv_admin",  ]
    redirect_unauthenticated_users = True
    raise_exception = True
    template_name = "conta/informe_rastreo_desecho.html"
    form_class = conta_f.RastreoDesechoInformeForm    

class InformeRastreoRepuesto(views.APIView):
    """ Lista todas los dispositivios que estan en desecho por medio de la `class:`Desecho` de inventario.
    """
    def get(self, request):      
        try:
            fecha_inicio = self.request.GET['fecha_min']
        except MultiValueDictKeyError:
            fecha_inicio=0
        
        try:
            fecha_fin = self.request.GET['fecha_max']
        except MultiValueDictKeyError:
            fecha_fin=0   
        try:
            tipo_dispositivo = [int(x) for x in self.request.GET.getlist('tipo_dispositivo[]')]
            if len(tipo_dispositivo) ==0:
                tipo_dispositivo = self.request.GET['tipo_dispositivo']            
        except MultiValueDictKeyError:
            tipo_dispositivo=0

        try:
            tipo_entrada = [int(x) for x in self.request.GET.getlist('tipo_entrada[]')]
            if len(tipo_entrada) ==0:
                tipo_entrada = self.request.GET['tipo_entrada']            
        except MultiValueDictKeyError:
            tipo_entrada=0
        try:
            entradas = [int(x) for x in self.request.GET.getlist('entrada_inventario[]')]
            list_entrada = [] 
            if len(entradas) ==0:
                entradas = self.request.GET['entrada_inventario'] 
                list_entrada.append(entradas)                          
        except MultiValueDictKeyError:
            entradas=0

        try:
            repuesto = [int(x) for x in self.request.GET.getlist('repuesto[]')]
            if len(repuesto) ==0:
                repuesto = self.request.GET['repuesto']            
        except MultiValueDictKeyError:
            repuesto=0

        try:
            estado_repuesto = [int(x) for x in self.request.GET.getlist('estado_repuesto[]')]
            if len(estado_repuesto) ==0:
                estado_repuesto = self.request.GET['estado_repuesto']            
        except MultiValueDictKeyError:
            estado_repuesto=0    
        sort_params ={}        
        crear_dict.crear_dict(sort_params,'tipo__id__in',tipo_dispositivo)
        crear_dict.crear_dict(sort_params,'entrada__fecha__gte',fecha_inicio)
        crear_dict.crear_dict(sort_params,'entrada__fecha__lte',fecha_fin)
        crear_dict.crear_dict(sort_params,'entrada__tipo__id__in',tipo_entrada)
        if len(list_entrada)==1:
            crear_dict.crear_dict(sort_params,'entrada__id__in',list_entrada)
        else:
            crear_dict.crear_dict(sort_params,'entrada__id__in',entradas)
        
        crear_dict.crear_dict(sort_params,'id__in',repuesto)
        crear_dict.crear_dict(sort_params,'estado__in',estado_repuesto) 
        repuesto_detalle = inv_m.Repuesto.objects.filter(**sort_params)
        datos_desecho =[]
        cantidad_total = 0   
        for data in repuesto_detalle:
            try:
                data_comentario = inv_m.RepuestoComentario.objects.filter(repuesto=data.id).last()
                comentario = data_comentario.comentario

            except:
                comentario = 0
            repuesto = {}
            repuesto["repuesto_id"] =data.id
            repuesto["repuesto_triage"] ="R-"+str(data.id)
            repuesto["repuesto_url"] =data.get_absolute_url()
            repuesto["repuesto_impreso"] =data.impreso
            repuesto["repuesto_tipo"] =data.tipo.tipo
            repuesto["repuesto_estado"] =data.estado.nombre
            repuesto["repuesto_descripcion"] =data.descripcion
            try:
                repuesto["repuesto_tarima"] =data.tarima.id
            except:
                repuesto["repuesto_tarima"]= 0
            repuesto["repuesto_valido"] =data.valido
            try:
                repuesto["repuesto_marca"] =data.marca.marca
            except:
                repuesto["repuesto_marca"] =0
            repuesto["repuesto_modelo"] =data.modelo
            repuesto["repuesto_creado"] =data.creada_por.get_full_name()
            repuesto["repuesto_entrada"] =data.entrada.id
            repuesto["repuesto_entrada_tipo"] =data.entrada.tipo.nombre
            repuesto["repuesto_entrada_url"] =data.entrada.get_absolute_url()
            repuesto["repuesto_entrada_fecha"] = str(data.entrada.fecha)
            repuesto["repuesto_desmembrado"] = inv_m.DispositivoRepuesto.objects.filter(repuesto=data.id).count()
            repuesto["repuesto_comentario"] = comentario
            repuesto["repuesto_total"] = repuesto_detalle.count() 
            datos_desecho.append(repuesto) 
        return Response(datos_desecho)
    
class RepuestoRastreoInformeView(LoginRequiredMixin, GroupRequiredMixin, FormView):
    """ Vista para obtener la informacion de los dispositivos para crear el informe de existencia mediante un
    api mediante el metodo GET  y lo muestra en el tempalte
    """
    group_required = [u"inv_conta",u"inv_admin", ]
    redirect_unauthenticated_users = True
    raise_exception = True
    template_name = "conta/informe_rastreo_repuesto.html"
    form_class = conta_f.RastreoRepuestoInformeForm   


"""
 Vista para la contabilidad del modulo de BEQT
"""

class ContabilidadBEQTEntradaInformeListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de PrecioEstandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarInformeForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/beqt/informe_entrada_beqt.html'
    form_class = conta_f.EntradaInformeForm

    def get_form(self, form_class=None):
        form = super(ContabilidadBEQTEntradaInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = beqt_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        form.fields['donante'].queryset = crm_m.Donante.objects.filter(nombre="BEQT")
        form.fields['tipo_entrada'].queryset = inv_m.EntradaTipo.objects.filter(nombre="BEQT")
        return form

"""
puntos de acceso de rest api para BEQT
"""
class InformeEntradaBeqtJson(views.APIView):
    """ Si es Donacion o Contenedor ir a traer el precio estandar del modelo de contabilidad
    en caso contrario desde el precio_unitario del detalle de entrada
    """
    def get(self, request):
        try:
            donante = self.request.GET['donante']
        except MultiValueDictKeyError as e:
            donante = 0
        try:
            tipo_entrada = []
            tipo_entrada = request.GET.getlist('tipo_entrada[]')
            if len(tipo_entrada) == 0:
                tipo = self.request.GET['tipo_entrada']
                tipo_entrada.append(tipo)
        except MultiValueDictKeyError as e:
            tipo_entrada = 0
        tipo_dispositivo = self.request.GET['tipo_dispositivo']
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        tipo_dispositivo_nombre = beqt_m.DispositivoTipoBeqt.objects.get(pk=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        # Obtener datos de Periodo Fiscal          
        lista = []

        # Preparar Filtros de Búsqueda
        q = []
        q.append(Q(tipo_dispositivo=tipo_dispositivo))

        if donante and donante != 0:
            q.append(Q(entrada__proveedor=donante))

        if tipo_entrada and tipo_entrada != 0:
            q.append(Q(entrada__tipo__in=tipo_entrada))

        # obtener Listado de Detalles de Entrada Aplicando Filtros
        entrada_detalle = beqt_m.EntradaDetalleBeqt.objects.values('entrada','precio_unitario','total').filter(
                    Q(fecha_dispositivo__gte=fecha_inicio),
                    Q(fecha_dispositivo__lte=fecha_fin),
                    reduce(AND,q)
                    ).exclude(entrada__tipo__nombre='Especial').annotate(Sum('precio_total'))


        # Obtener Existencia Inicial y Saldo Inicial
        fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)        
        totales_anterior = get_existencia_beqt(tipo_dispositivo,fecha_inicial)
        #precio_anterior = totales_anterior['precio_estandar']
        precio_total_anterior = totales_anterior['saldo_total']
        existencia_anterior = totales_anterior['existencia']

        # Obtener Existencia Final y Saldo Final
        total_actual = get_existencia_beqt(tipo_dispositivo,fecha_fin)
        #precio_actual = total_actual['precio_estandar']
        precio_total = total_actual['saldo_total']
        existencia_actual = total_actual['existencia']
        for datos_entrada in entrada_detalle:           
            entrada = beqt_m.Entrada.objects.get(pk=datos_entrada['entrada'])
            precio = datos_entrada['precio_unitario']
            cantidad = datos_entrada['total']
            dispositivo = {}
            # Validar Precio de Compra y Donación
            if (not precio or precio == 0) and not entrada.tipo.contable:
                # Obtener Precio Estandar
                pass 
                
            else:
                #precio = 0                    
                print("Ingreso aca 2")
            dispositivo['fecha'] = entrada.fecha
            dispositivo['id'] = entrada.id
            dispositivo['url'] = entrada.get_absolute_url()
            dispositivo['util'] = cantidad
            dispositivo['precio'] = precio
            dispositivo['tipo'] = entrada.tipo.nombre
            dispositivo['proveedor'] = entrada.proveedor.nombre
            try:
                dispositivo['total'] = cantidad * precio                
            except:
                print(entrada.id)
            dispositivo['total_costo'] = precio_total_anterior
            dispositivo['total_final'] = existencia_anterior
            dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
            dispositivo['tipo_dispositivo'] = tipo_dispositivo_nombre.tipo
            dispositivo['total_costo_despues'] = precio_total
            dispositivo['total_despues'] = existencia_actual
            lista.append(dispositivo)
        return Response(lista)
    
class ContabilidadEntradaDispBeqtInformeListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de Dispositivos x Entrada mediante una :class:`PrecioEstandar`
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/beqt/informe_entrada_dispositivo.html'
    form_class = conta_f.EntradaDispositivoBeqtInformeForm

    def get_form(self, form_class=None):
        form = super(ContabilidadEntradaDispBeqtInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = beqt_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form

class InformeEntradaDispositivoBeqtJson(views.APIView):
    """ Lista los dispositivos ingresados al sistema por rango de fechas
    """
    def get(self, request):
        try:
            no_entrada = self.request.GET['no_entrada']
        except MultiValueDictKeyError as e:
            no_entrada = 0

        tipo_dispositivo = self.request.GET['tipo_dispositivo']
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        tipo_dispositivo_nombre = beqt_m.DispositivoTipoBeqt.objects.get(pk=tipo_dispositivo)       
        

        # Obtener datos de Periodo Fiscal       
        lista_dispositivos = {}
        lista = []

        # Preparar Filtros de Búsqueda
        q = []
        q.append(Q(tipo=tipo_dispositivo))

        if no_entrada and no_entrada != 0:
            q.append(Q(entrada_detalle__entrada=no_entrada))

        # obtener Listado de Detalles de Entrada Aplicando Filtros
        entrada_detalle = beqt_m.DispositivoBeqt.objects.filter(
                    Q(entrada_detalle__fecha_dispositivo__gte=fecha_inicio),
                    Q(entrada_detalle__fecha_dispositivo__lte=fecha_fin),
                    reduce(AND,q)
                    ).exclude(entrada_detalle__entrada__tipo__nombre='Especial').order_by('triage')

        
        nuevo_dispositivo = []
        nuevo_dispositivo = sorted(entrada_detalle,key=lambda s:int(re.search(r'\d+',s.triage).group()))
        # Obtener Existencia Inicial y Saldo Inicial
        fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
        totales_anterior = get_existencia_beqt(tipo_dispositivo,fecha_inicial)
        existencia_anterior = totales_anterior['existencia']

        # Obtener Existencia Final y Saldo Final
        total_actual = get_existencia_beqt(tipo_dispositivo,fecha_fin)
        existencia_actual = total_actual['existencia']

        for detalle in nuevo_dispositivo:
            entrada = beqt_m.Entrada.objects.get(pk=detalle.entrada_detalle.entrada.id)
            dispositivo = {}

            dispositivo['triage'] = detalle.triage
            dispositivo['entrada'] = detalle.entrada_detalle.entrada.id
            dispositivo['fecha'] = detalle.entrada_detalle.fecha_dispositivo
            dispositivo['tipo_entrada'] = detalle.entrada_detalle.entrada.tipo.nombre
            dispositivo['url'] = detalle.get_absolute_url()
            dispositivo['url_entrada'] = detalle.entrada_detalle.entrada.get_absolute_url()

            dispositivo['total_final'] = existencia_anterior
            dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
            dispositivo['tipo_dispositivo'] = tipo_dispositivo_nombre.tipo
            dispositivo['total_despues'] = existencia_actual
            lista.append(dispositivo)
        return Response(lista)
    

    
class ContabilidadBeqtSalidasInformeListView(LoginRequiredMixin, FormView):
    """Vista utilizada para listar las salidas por rango de fechas y tipo de dispositivo.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/beqt/informe_salidas.html'
    form_class = conta_f.SalidasBeqtInformeForm
    def get_form(self, form_class=None):
        form = super(ContabilidadBeqtSalidasInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = beqt_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form
       
class InformeSalidaBeqtJson(views.APIView):
    """ Listar las salidas por el tipo seleccionado, si es compra o no traer o excluir los dispositivos de compra del listado.
    """
    def get(self, request):
        try:
            udi = self.request.GET['udi']
        except MultiValueDictKeyError as e:
            udi = 0

        try:
            beneficiado = self.request.GET['beneficiado']
        except MultiValueDictKeyError as e:
            beneficiado = 0        

        try:
            tipo_salida = []
            tipo_salida = request.GET.getlist('tipo_salida[]')
            if len(tipo_salida) == 0:
                tipo = self.request.GET['tipo_salida']
                tipo_salida.append(tipo)
        except MultiValueDictKeyError as e:
            tipo_salida = 0

        tipo_dispositivo = self.request.GET['tipo_dispositivo']
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        tipo_dispositivo_nombre = beqt_m.DispositivoTipoBeqt.objects.get(pk=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        # Obtener datos de Periodo Fiscal
       
        salida_especial = inv_m.SalidaTipo.objects.get(especial=True)
        #print(inv_m.EntradaTipo.objects.filter(contable=True).last())
        tipo_compra =  inv_m.EntradaTipo.objects.filter(contable=True).first()
        lista_dispositivos = {}
        lista = []

        # Armar Query de Consulta
        sql_select = '''SELECT isi.id as 'Salida', COUNT(*) AS 'CANTIDAD'
                FROM beqt_dispositivopaquete idp
                INNER JOIN beqt_dispositivobeqt id on idp.dispositivo_id = id.id
                INNER JOIN beqt_paquetebeqt ip on idp.paquete_id = ip.id
                INNER JOIN beqt_salidainventario isi on ip.salida_id = isi.id
                INNER JOIN beqt_entrada ie on id.entrada_id = ie.id
                '''

        sql_where = """ WHERE id.tipo_id = {tipo_dispositivo}
                AND isi.en_creacion = 0
                AND isi.fecha between '{fecha_inicio}' AND '{fecha_fin}'
                AND isi.tipo_salida_id <> {especial}
                """.format(tipo_dispositivo = tipo_dispositivo,
                                                                                    fecha_inicio=fecha_inicio,
                                                                                    fecha_fin=fecha_fin,
                                                                                    especial=salida_especial.id)
        sql_group = " GROUP BY isi.id"

        if udi and udi != 0:
            escuela = escuela_m.Escuela.objects.get(codigo=udi)
            if escuela:
                sql_where += " AND isi.escuela_id = " + str(escuela.id)

        if beneficiado and beneficiado != 0:
            sql_where += " AND isi.beneficiario_id = " + str(beneficiado)

        if tipo_salida and tipo_salida != 0:
            sql_where += " AND isi.tipo_salida_id in (" + ','.join(tipo_salida) + ")"

        """if compra and not donaciones:
            sql_where += " AND ie.tipo_id = " + str(tipo_compra.id)
        elif not compra and donaciones:
            sql_where += " AND ie.tipo_id <> " + str(tipo_compra.id)"""

        sql_query = sql_select + sql_where + sql_group

        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            result =cursor.fetchall()

        # Obtener Existencia Inicial y Saldo Inicial
        fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
        totales_anterior = get_existencia_beqt(tipo_dispositivo,fecha_inicial)
        #precio_anterior = totales_anterior['precio_estandar']
        precio_total_anterior = totales_anterior['saldo_total']
        existencia_anterior = totales_anterior['existencia']

        # Obtener Existencia Final y Saldo Final
        total_actual = get_existencia_beqt(tipo_dispositivo,fecha_fin)
        #precio_actual = total_actual['precio_estandar']
        precio_total = total_actual['saldo_total']
        existencia_actual = total_actual['existencia']        
        for datos_salida in result:
            #print(datos_salida)
            salida = beqt_m.SalidaInventario.objects.get(pk=datos_salida[0])
            cantidad = datos_salida[1]
            precio = 1
            total = cantidad * 1
            dispositivo = {}

            if salida.escuela:
                beneficiado = salida.escuela.nombre
            elif salida.beneficiario:
                beneficiado = salida.beneficiario.nombre
            else:
                beneficiado = ''

            dispositivo['fecha'] = salida.fecha
            dispositivo['id'] = salida.id
            dispositivo['no_salida'] = salida.no_salida
            dispositivo['url'] = salida.get_absolute_url()
            dispositivo['util'] = cantidad
            dispositivo['precio'] = precio
            dispositivo['tipo'] = salida.tipo_salida.nombre
            dispositivo['beneficiado'] = beneficiado
            dispositivo['total'] = total
            dispositivo['total_costo'] = precio_total_anterior
            dispositivo['total_final'] = existencia_anterior
            dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
            dispositivo['tipo_dispositivo'] = tipo_dispositivo_nombre.tipo
            dispositivo['total_costo_despues'] = precio_total
            dispositivo['total_despues'] = existencia_actual
            lista.append(dispositivo)
        return Response(lista)
    

class ContabilidadResumenBeqtInformeListView(LoginRequiredMixin, FormView):
    """Vista utilizada para listar el resumen de Inventario.
    """
    model = conta_m.PrecioEstandar
    template_name = 'conta/beqt/informe_resumen.html'
    form_class = conta_f.ResumenInformeBeqtForm

    def get_form(self, form_class=None):
        form = super(ContabilidadResumenBeqtInformeListView, self).get_form(form_class)
        form.fields['tipo_dispositivo'].queryset = beqt_m.AsignacionTecnico.objects.get(usuario=self.request.user ).tipos.filter(usa_triage=True)
        return form
    


class InformeResumenBeqtJson(views.APIView):
    """ Lista todas las salidas de desecho con triage que han sucedido en un rango de fechas,
    Solamente cuentan aquellas salidas que han sido cerradas.
    """
    def get(self, request):
        fecha_inicio = self.request.GET['fecha_min']
        fecha_fin = self.request.GET['fecha_max']
        try:
            tipo_dispositivo = []
            tipo_dispositivo = self.request.GET.getlist('tipo_dispositivo[]')
            if len(tipo_dispositivo) == 0:
                tipo = self.request.GET['tipo_dispositivo']
                tipo_dispositivo.append(tipo)
        except MultiValueDictKeyError as e:
            tipo_dispositivo = 0

        # Filtrar por tipos de dispositivos seleccionados
        #print(self.request.user.tipos_dispositivos.tipos.filter())
        if tipo_dispositivo == 0 or not tipo_dispositivo:
            dispositivos = self.request.user.tipos_dispositivos_beqt.tipos.all()
        else:
            dispositivos = self.request.user.tipos_dispositivos_beqt.tipos.filter(id__in=tipo_dispositivo)

        # Validar que el rango de fechas pertenezcan a un solo período fiscal
        validar_fecha = conta_m.PeriodoFiscal.objects.filter(fecha_inicio__lte=fecha_inicio, fecha_fin__gte=fecha_fin)
        acumulador = 0
        acumulador_anterior = 0
        acumulador_ant_ex = 0
        acumulador_act_ex = 0

        if validar_fecha.count() == 1:
            # Obtener datos de Periodo Fiscal
            periodo = validar_fecha[0]
            lista_dispositivos = {}
            lista = []
            for tipo in dispositivos:               
                dispositivo = {}
                # Obtener Saldo Anterior
                fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d') - timedelta(days=1)
                totales_anterior = get_existencia_beqt(tipo,fecha_inicial)
                precio_total_anterior = totales_anterior['saldo_total']
                existencia_anterior = totales_anterior['existencia']
                acumulador_anterior += precio_total_anterior
                acumulador_ant_ex += existencia_anterior

                # Obtener Saldo Actual
                total_actual = get_existencia_beqt(tipo,fecha_fin)
                precio = 1
                precio_total = total_actual['saldo_total']
                existencia = total_actual['existencia']
                acumulador += precio_total
                acumulador_act_ex += existencia
                # print("Tipo >{} , Precio > {}",tipo,precio)
                # Obtener Total de Entradas
                entradas = beqt_m.EntradaDetalleBeqt.objects.filter(
                        Q(fecha_dispositivo__gte=fecha_inicio),
                        Q(fecha_dispositivo__lte=fecha_fin),
                        Q(tipo_dispositivo=tipo)
                        ).aggregate(Sum('total'))

                if entradas['total__sum'] is not None:
                    entradas = entradas['total__sum']
                else:
                    entradas = 0

                # Obtener Total de Salidas
                salida_especial = inv_m.SalidaTipo.objects.get(especial=True)
                salidas = len(beqt_m.DispositivoPaquete.objects.filter(
                        Q(paquete__salida__fecha__gte=fecha_inicio),
                        Q(paquete__salida__fecha__lte=fecha_fin),
                        Q(dispositivo__tipo=tipo),
                        Q(paquete__salida__en_creacion=False)))

                """desecho = len(inv_m.DesechoDispositivo.objects.filter(
                        Q(desecho__fecha__gte=fecha_inicio),
                        Q(desecho__fecha__lte=fecha_fin),
                        Q(dispositivo__tipo=tipo),
                        Q(desecho__en_creacion=False)))"""

                #salidas += desecho

                dispositivo['tipo'] = tipo.tipo
                dispositivo['existencia_anterior'] = existencia_anterior
                dispositivo['saldo_anterior'] = precio_total_anterior
                dispositivo['entradas'] = str(entradas)
                dispositivo['salidas'] = salidas
                dispositivo['existencia'] = existencia
                dispositivo['saldo_actual'] = precio_total
                dispositivo['costo_inicial'] = acumulador_anterior
                dispositivo['total_inicial'] = acumulador_ant_ex
                dispositivo['rango_fechas'] = str(fecha_inicio)+"  AL  "+str(fecha_fin)
                dispositivo['costo_final'] = acumulador
                dispositivo['total_final'] = acumulador_act_ex
                lista.append(dispositivo)
            return Response(lista)
        else:
            print("No existe en el periodo fiscal")


class PlanillaApiConta(LoginRequiredMixin,views.APIView): 
    """
    Vista que maneja la carga del archivo xls
    """
    def post(self, request):
        """
        Método POST de carga de archivo xlsx y su almacenamiento
        """
        form=conta_f.PlanillaForm()
        if request.method=='POST' and request.FILES.get("myfile"):
            """
            Verifíca que sea un método post al mismo tiempo que dentro del request se haya capturado un file
            """
            excel_path=os.path.join(settings.MEDIA_ROOT,"planilla","excel")
            file_list =os.listdir(excel_path)
            number_file=len(file_list)
            myfile=request.FILES["myfile"]
            new_name=str('PlanillaN') + str(number_file) +str(".")+str(myfile.name.split(".")[-1])          
            file_storage=FileSystemStorage(location=excel_path)
            file_name=file_storage.save(new_name, myfile)  
            url=str(settings.MEDIA_URL)+"planilla/output/Comprobantes.zip"
            context={
                'form':form,
                'file_name':file_name,
                'url':url,
                'file_cargado':str(myfile),
                'titulo':"Descargar comprobantes en zip"
            }              
            return render(
                request,
                'conta/planilla/planilla.html',
                 context
                    )
        return render(
                request,
                'conta/planilla/planilla.html'
            )  
             
                  
class Empleado(object):
    """
    Clase que que contiene y modela los distintos tipos de planillas
    """
    nombre = ''
    codigo = ''
    correo = ''
    cuenta = ''
    primera = 0.0
    decreto = 0.0
    descuentos = 0.0
    asdeco = 0.0
    igss = 0.0
    salario = 0.0
    total_descuentos = 0.0
    depositar = 0.0

    def __init__(self, quincena=1, datos=[], *args, **kwargs):
        if quincena == 1:
            self.crear_primera(datos)
        elif quincena == 2:
            self.crear_segunda(datos)
        elif quincena == 3:
            self.crear_bono(datos)
        elif quincena == 4:
            self.crear_aguinaldo(datos)

    def __repr__(self):
        return self.nombre

    def crear_primera(self, datos):
        """
        Método que configura los valores necesarios para procesar la planilla en la primera quincena
        """
        self.nombre = datos[1].value
        self.codigo = datos[0].value
        self.correo = datos[2].value
        self.cuenta = datos[5].value
        self.primera = float(datos[6].value)
        self.decreto = float(datos[7].value)
        self.descuentos = float(datos[9].value)
        self.igss = float(datos[10].value)
        self.asdeco = float(datos[11].value)
        self.salario = self.primera + self.decreto
        self.total_descuentos = self.descuentos + self.asdeco + self.igss
        self.depositar = round(float(self.salario - self.total_descuentos), 2)
        

    def crear_segunda(self, datos):
        """
        Método que configura los valores necesarios para procesar la planilla en la segunda quincena
        """
        self.nombre = datos[1].value
        self.codigo = datos[0].value
        self.correo = datos[2].value
        self.cuenta = datos[4].value
        self.sueldo = float(datos[5].value)
        self.bonificacion = float(datos[6].value)
        self.otros = float(datos[7].value)
        self.quincena = float(datos[9].value)
        self.igss2 = float(datos[10].value)
        self.isr = float(datos[11].value)
        self.boleto = float(datos[12].value)
        self.asdeco2 = float(datos[13].value)
        self.descuentos = float(datos[14].value)
        self.devengado = self.sueldo + self.bonificacion + self.otros
        self.total_descontado = self.quincena + self.igss2 + self.isr + self.boleto + self.asdeco2 + self.descuentos
        self.depositar2 = self.devengado - self.total_descontado

    def crear_bono(self, datos):
        """
        Método que configura los valores necesarios para procesar la planilla de tipo Bono 14
        """
        self.nombre = datos[1].value
        self.codigo = datos[0].value
        self.correo = datos[2].value
        self.cuenta = datos[3].value
        self.bono = round(datos[20].value, 2)
        self.letras = num2words(self.bono, lang='es')

    def crear_aguinaldo(self, datos):
        """
        Método que configura los valores necesarios para procesar la planilla de tipo Aguinaldo
        """
        self.nombre = datos[1].value
        self.codigo = datos[0].value
        self.correo = datos[2].value
        self.cuenta = datos[3].value
        self.aguinaldo = round(datos[20].value, 2)
        self.letras = num2words(self.aguinaldo, lang='es')



class PlanillaParse:
    """
    Clase con variedad de métodos capaces de parsear los datos leídos de un excel para generar comprobantes de planillas
    """
    lista_empleados = []
    fecha_quincena = ''

    def __init__(self, file_name,planilla):  
        self.file_name = file_name
        self.planilla=int(planilla)
        self.parse_Planilla()
    
    def load_Planilla(self,workbook,name_planilla,date_extra_rows,data_extra_rows):
        """
        Método encargado de recorrer el archivo leído, agruparlo y proceder a la creación del setteo de datos por tipo de planilla
        """
        fila_inicial=1   
        for row in workbook[name_planilla].iter_rows(min_row=1):
            if any(row):
                break
            fila_inicial=+1 
        self.fecha_quincena = workbook[name_planilla][(fila_inicial+date_extra_rows)][0].value
        for _, row in enumerate(workbook[name_planilla].iter_rows(min_row=(fila_inicial+data_extra_rows))):
                
            if any(row) and row[0].value is not None:
                self.lista_empleados.append(Empleado(quincena=self.planilla,datos=row))
        

    def parse_Planilla(self):
        """
        Método por el cual se controla que tipo de planilla (
        1:Primera quincena
        2:Segunda quincena, 
        3:Bono 14, 
        4:Aguinaldo, 
        )  se van a generar según el archivo obtenido
        """
        self.lista_empleados=[] 
        workbook=None 
        try:
            workbook=load_workbook(self.file_name, data_only=True)
        except Exception as e:
            print(e)
      
        if self.planilla == 1:
            self.load_Planilla(workbook,'Anticipo',2,6)
        elif self.planilla == 2:
            self.load_Planilla(workbook,'Fin de Mes',2,6)
        elif self.planilla == 3 or self.planilla == 4:
            self.load_Planilla(workbook,'Nomina',3,6)
            
        return self.lista_empleados
    

class html_generadoenerator:
    """
    Clase creada con la finalidad de la creación de html, pdf y otros archivos de salida
    """
    options = {
        'encoding': "UTF-8",
        'enable-local-file-access': '',
        'margin-top': '0.7in',
        'margin-right': '0.1in',
        'margin-bottom': '1.5in',
        'margin-left': '0.6in',
    }
    @classmethod
    def configurarFiltros(cls, disposición,tipo_planilla):
        """Método de clase que al obtener los valores númericos del tipo y disposición de planilla retorna
        una tupla con las cadenas representativas de dichos valores de entrada
        """
        tipos_planilla=['primera','segunda','bono','aguinaldo']
        tipo_planillaTexto=tipos_planilla[int(tipo_planilla)-1]
        if disposición == 1:
            return tipo_planillaTexto,'unico'
        else:
            return tipo_planillaTexto,'lista'
        
    def __init__(self,template_planilla,template_tipo,enviar , output_folder='html/output', template_folder='html/templates',destinatario=' '):
        self.template_name = 'template_{}_{}.html'.format(template_planilla, template_tipo)
        self.template_folderAux=template_folder
        self.output_folderAux=output_folder
        self.output_folder=pathlib.Path(self.output_folderAux)
        self.template_folder=pathlib.Path(self.template_folderAux)
        self.env=Environment(loader=FileSystemLoader(str(self.template_folder)))       
        self.output_folder=pathlib.Path(self.output_folder)
        self.enviar=enviar
        self.destinatario=destinatario
        
        
       
    @classmethod
    def copy(cls, src, dest):
        """
        Método para copiar un directorio desde una locación actual a una destino
        """
        try:
            print("Se guarda en: ")
            shutil.copytree(src, dest)
        except OSError as e:
            if e.errno == errno.ENOTDIR:
                print(shutil.copy(src, dest))
            else:
                pass    

    def crear_lista(self,planilla):
        """
        Método que genera la disposición de lista de los comprobantes de planilla para los empleados
        """
        
        ruta_logo = str(pathlib.Path(settings.BASE_DIR) / 'static/image/Logo_Funsepa-01.png').replace(os.sep,"/")
        ruta_qr = str(pathlib.Path(settings.BASE_DIR) / 'static/image/qr_contador.png').replace(os.sep,"/")
        logo_path = "file:///"+ruta_logo
        qr_path = "file:///"+ruta_qr
        template=self.env.get_template(self.template_name)
        full_name_template=str(self.output_folder/'lista_empleados.html')
        html_output=template.render(
            lista_empleados=planilla.lista_empleados,
            quincena=planilla.fecha_quincena,
            hora=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            qr=qr_path,
            logo=logo_path
        )
        with open(full_name_template,"w",encoding='utf-8') as html:
            html.write(html_output)
        #Generar pdf
        pdf=self.generar_pdf(full_name_template,"lista_empleados")
        pdf_list=[]
        if self.enviar and pdf is not None:
            flag=self.send_mail(pdf,planilla.fecha_quincena, EMAIL_CREDENCIALES['from_contabilidad'])
            if flag:
                ##Envio de correo exitosamente
                return 0,''
            else:
                ##Envio de correo no exitoso
                return -1,''
        else:
            ##No envío de correo
            pdf_list.append(str(pdf+'.pdf'))
            return 1,pdf_list
            

    def crear_unico(self,planilla):
        """
        Método que crea la disposición de los comprobanes de planilla de forma individual por cada empleado
        """
        template=self.env.get_template(self.template_name)
        retorno=-1
        pdf_list=[]
        ruta_logo = str(pathlib.Path(settings.BASE_DIR) / 'static/image/Logo_Funsepa-01.png').replace(os.sep,"/")
        ruta_qr = str(pathlib.Path(settings.BASE_DIR) / 'static/image/qr_contador.png').replace(os.sep,"/")
        logo_path = "file:///"+ruta_logo
        qr_path = "file:///"+ruta_qr
        for empleado in planilla.lista_empleados:
            full_name_template=str(self.output_folder/'{}.html'.format(empleado.codigo))
            html_output=template.render(
            persona=empleado,
            quincena=planilla.fecha_quincena,
            hora=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            qr=qr_path,
            logo=logo_path
            )
            with open(full_name_template,"w",encoding='utf-8') as html:
                html.write(html_output)
            #Generar pdf
            pdf=self.generar_pdf(full_name_template,empleado.codigo)
            if self.enviar and pdf is not None:
                flag=self.send_mail(pdf,planilla.fecha_quincena,empleado.correo)
                if flag:
                    ##Envio de correo exitosamente
                    retorno= 0
                else:
                    ##Envio de correo no exitoso
                    retorno -1
            else:
                ##No envio de correo
                pdf_list.append(str(pdf+'.pdf'))
                retorno=1
        return retorno,pdf_list
                
    def empaquetar_pdf(self,pdf_list):
        """
        Método que comprime el(los) comprobante(s) creados para la disposición y tipo de planilla seleccionada
        """
        try:
            zip_file="Comprobantes.zip"
            path_zip=str(self.output_folder/zip_file)
            with zipfile.ZipFile(path_zip,'w') as zipc:
                for pdf in pdf_list:
                    if os.path.exists(pdf):
                        auxpdf=pdf.split("\\")
                        pdfile_path=auxpdf[len(auxpdf)-1]
                        zipc.write(pdf,os.path.basename(pdfile_path))
            return True
        except Exception as e:
            print(e)
            return Response(
                "Ha ocurrido un error en la descarga de los comprobantes",
                status=500
            )
        

    def generar_pdf(self,template,fname):
        """
        Método que genera los pdf requeridos utilizando archivos html para luego guardarlos
        """
        actual_so=platform.system()
        ruta_wkhmtltopdf= shutil.which("wkhtmltopdf")
        if ruta_wkhmtltopdf is None:
            if actual_so=="nt":

                ruta_wkhmtltopdf="C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe"
            elif actual_so=="posix":
                ruta_wkhmtltopdf = "/usr/local/bin/wkhtmltopdf"
            elif actual_so=="Windows":
                ruta_wkhmtltopdf = "C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe"
            elif actual_so=="Linux":
                #ruta_wkhmtltopdf = ""
                ruta_wkhmtltopdf = "/usr/local/bin/wkhtmltopdf"                
            else:
                raise Exception("No se ha podido encontrar el Sistema Operativo")
        config=pdfkit.configuration(wkhtmltopdf=ruta_wkhmtltopdf)
        file_name=str(self.output_folder/fname)
        try:
            pdfkit.from_file(
            template,
            '{}.pdf'.format(file_name),
            options=self.options,
            configuration=config
            )
        except Exception as e:
            print(str(e))
        return file_name
    
    def send_mail(self,pdf,quincena,to):
        """
        Método que envía los correos electrónicos los cuales contienen los comprobantes como archivo adjunto
        """
        try:
            content ={
                'subject': "Boleta de Pago",
                'body': """Boleta de pago: {}.
        Este mensaje fue generado de forma automática por SUNI por favor no contestar este correo.
        Para dudas o comentarios favor escribir a {} o escanee el código QR en el documento""".format(quincena, EMAIL_CREDENCIALES['from_contabilidad'])
            }
            email=mail.EmailMessage(
            content['subject'],
            content['body'],
            'soporte_contabilidad@funsepa.org',
            [to],
            connection=mail.get_connection(
                backend=EMAIL_CREDENCIALES['backend'],
                username=EMAIL_CREDENCIALES['username'],
                password=EMAIL_CREDENCIALES['password'],
                host=EMAIL_CREDENCIALES['host'],
                port=587,
                use_tls=True,
                fail_silently=False
                )
            )
            email.attach_file(str(pdf+".pdf"))
            email.send()
            return True
            
        except Exception as e:
            print(e)
            return False


    def delete_files(self,excel_path):
        """
        Método que elimina los archivos de salida a su totalidad al ser enviado un correo, borra html, pdf y comprimidos en zip
        """
        for archivos in os.listdir(excel_path):
            path_archivo=os.path.join(excel_path,archivos)
            if os.path.isfile(path_archivo):
                os.remove(path_archivo)
        for archivos in os.listdir(str(self.output_folder)):
            path_archivo=os.path.join(str(self.output_folder),archivos)
            if os.path.isfile(path_archivo):
                os.remove(path_archivo)
    
    

class FiltrosPlanillaApiConta(LoginRequiredMixin,views.APIView): 
    """
    Este genera los html y los pdf
    """
    def post(self, request):            
            if request.method=='POST':
                    planilla_seleccionada=self.request.POST['tipo_de_planilla']
                    distribucion_seleccionada=self.request.POST['disposicion_planilla']
                    correo=request.POST.get('enviar_correo')
                    excel_path=os.path.join(settings.MEDIA_ROOT,'planilla','excel')
                    send_mail=False
                    if correo: 
                        send_mail=True
                    file_path=os.listdir(excel_path)[len(os.listdir(excel_path))-1]
                    file_name=os.path.join(excel_path,file_path)             
                    datos = PlanillaParse(file_name,planilla_seleccionada)
                    ##Manejo de html's y pdf's                    
                    output_folder=os.path.join(settings.MEDIA_ROOT,'planilla','output')
                    template_folder=os.path.join(settings.BASE_DIR,'templates/conta/planilla/')
                    tipo,disposicion=html_generadoenerator.configurarFiltros(int(distribucion_seleccionada),int(planilla_seleccionada))                    
                    html_generado=html_generadoenerator(template_planilla=tipo,
                                                template_tipo=disposicion,
                                                output_folder=output_folder,
                                                enviar=send_mail, 
                                                template_folder=template_folder
                                                )
                    flag=1
                    pdf=[]
                    if disposicion=='unico':
                       flag,pdf=html_generado.crear_unico(datos)
                    else:
                        flag,pdf=html_generado.crear_lista(datos)
                    if send_mail and flag==0:
                        html_generado.delete_files(excel_path=os.path.join(settings.MEDIA_ROOT,'planilla','excel'))
                        return JsonResponse({"success":"Se han enviado los comprobantes de pago y eliminado los archivos exitosamente"},status=200)
                    elif send_mail==False and flag==1:
                        flag= html_generado.empaquetar_pdf(pdf)
                        path_url=os.path.join(settings.MEDIA_URL,"planilla/output/Comprobantes.zip")
                        return JsonResponse({"empaquetado":"{}".format(path_url)},status=200)
                        
                    elif send_mail and flag==-1:
                        return JsonResponse({"error":"No se han enviado comprobantes, compruebe de nuevo"},status=200)
                    
              
    
class PlanillaContaView(LoginRequiredMixin,GroupRequiredMixin,FormView):
    """Vista que maneja el template de Planillas de Contabilidad"""
    template_name = 'conta/planilla/planilla.html'
    form_class = conta_f.PlanillaForm
    group_required = [u"conta_planilla", ]
    def get_context_data(self, **kwargs):
        context = super(PlanillaContaView, self).get_context_data(**kwargs)   
        context['url']=os.path.join(settings.MEDIA_URL,"planilla/output/Comprobantes.zip")
        return context

#Fin de la planilla de contabilidad

class RastreoDispositivoContabilidad(views.APIView):
    """ Lista todas los dispositivos  con triage que han sucedido en un rango de fechas,
    entradas, salidas, proyectos y mas.
    """
    def get(self, request):
        try:
            fecha_inicio = self.request.GET['fecha_min']
        except:
            fecha_inicio =0
        try:
            fecha_fin = self.request.GET['fecha_max']
        except:
            fecha_fin = 0
        try:    
            triage = self.request.GET['triage']
        except:
            triage = 0
        try:
            entrada = self.request.GET['no_entrada']
        except:
            entrada = 0
        try:
            salida = self.request.GET['salida']
        except:
            salida = 0
        lista = []
        try:
            tipo_dispositivo = []
            tipo_dispositivo = self.request.GET.getlist('tipo_dispositivo[]')
            if len(tipo_dispositivo) == 0:
                tipo = self.request.GET['tipo_dispositivo']
                tipo_dispositivo.append(tipo)
        except MultiValueDictKeyError as e:
            tipo_dispositivo = 0
        
        try:
            proyecto = []
            proyecto = self.request.GET.getlist('proyecto[]')
            if len(proyecto) == 0:
                proyecto_solo = self.request.GET['proyecto']
                proyecto.append(proyecto_solo)
        except MultiValueDictKeyError as e:
            proyecto = 0
        sort_params = {}
        if salida == 0:
            crear_dict.crear_dict(sort_params,'dispositivo__entrada_detalle__proyecto__in',proyecto)
            crear_dict.crear_dict(sort_params,'fecha__lte',fecha_fin)
            crear_dict.crear_dict(sort_params,'fecha__gte',fecha_inicio)
            crear_dict.crear_dict(sort_params,'dispositivo__triage',triage)
            crear_dict.crear_dict(sort_params,'dispositivo__entrada',entrada)
            crear_dict.crear_dict(sort_params,'dispositivo__tipo__id__in',tipo_dispositivo)
            crear_dict.crear_dict(sort_params,'tipo_movimiento',1) 
            data  = conta_m.MovimientoDispositivo.objects.filter(**sort_params)
        else:
            crear_dict.crear_dict(sort_params,'dispositivo__entrada_detalle__proyecto__in',proyecto)
            crear_dict.crear_dict(sort_params,'fecha__lte',fecha_fin)
            crear_dict.crear_dict(sort_params,'fecha__gte',fecha_inicio)
            crear_dict.crear_dict(sort_params,'dispositivo__triage',triage)
            crear_dict.crear_dict(sort_params,'dispositivo__entrada',entrada)
            crear_dict.crear_dict(sort_params,'referencia__icontains',salida)
            crear_dict.crear_dict(sort_params,'dispositivo__tipo__id__in',tipo_dispositivo)
            crear_dict.crear_dict(sort_params,'tipo_movimiento',-1) 
            data  = conta_m.MovimientoDispositivo.objects.filter(**sort_params)       
        for info in data:
            data_resultado= info.dispositivo.rastreo_dispositivo()                       
            lista.append(data_resultado)        
        return Response(lista)
    
class InformeRastreoDispositivoView(LoginRequiredMixin, FormView):
    """ Vista para obtener la informacion de los dispositivos para crear el informe de rastreo mediante un
    api y el metodo GET  y lo muestra en el template
    """
    redirect_unauthenticated_users = True
    raise_exception = True
    template_name = "conta/informe_rastreo_dispositivo.html"
    form_class = conta_f.RastreoDispositivoInformeForm