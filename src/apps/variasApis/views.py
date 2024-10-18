from django.shortcuts import render
from apps.inventario import models as inv_m
from apps.cyd import models as cyd_m
from apps.main import models as main_m 
from apps.escuela import models as esc_m 
from apps.tpe import models as tpe_m 
import qrcode
import json
from io import BytesIO
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook
import os
from django.conf import settings
from django.contrib.auth.models import User
from apps.escuela.models import *
import json
from apps.main.models import Municipio
from rest_framework import views, status
from rest_framework.response import Response
import os.path as path
import os
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from openpyxl import load_workbook
from apps.mye.models  import Cooperante
from apps.Evaluacion.models import estadoFormulario, AsignacionPregunta, Formulario

# Create your views here.
class SubirTodo(views.APIView): 
    """  Clase encargada de Subir todos los datos de los archivos json a las tablas del modulo de cyd del SUNI
    """
    def get(self, request):
            ruta = "C:/Users/Edgar/Documents/TablasMigrarSuni/json/"
            file_escuela = open(str(ruta)+"cyd_escuela_new.json",encoding="utf-8")
            file_asesoria = open(str(ruta)+"cyd_asesoria.json",encoding="utf-8")
            file_asignacion = open(str(ruta)+"cyd_asignacion.json",encoding="utf-8")#1
            file_calendario = open(str(ruta)+"cyd_calendario.json",encoding="utf-8")#2
            file_crasistencia = open(str(ruta)+"cyd_crasistencia.json",encoding="utf-8")#3
            file_crhito = open(str(ruta)+"cyd_crhito.json",encoding="utf-8")#4
            file_curso = open(str(ruta)+"cyd_curso.json",encoding="utf-8")#5
            file_grupo = open( str(ruta)+"cyd_grupo.json",encoding="utf-8")#6
            file_notaasistencia = open( str(ruta)+"cyd_notaasistencia.json",encoding="utf-8")#7
            file_notahito = open(str(ruta)+"cyd_notahito.json",encoding="utf-8")#8
            file_parescolaridad =open(str(ruta)+"cyd_parescolaridad.json",encoding="utf-8")#9
            file_paretnia = open( str(ruta)+"cyd_paretnia.json",encoding="utf-8")#10
            file_pargenero = open( str(ruta)+"cyd_pargenero.json",encoding="utf-8")#11
            file_parrol = open( str(ruta)+"cyd_parrol.json",encoding="utf-8")#12         
            file_participante =open( str(ruta)+"cyd_participante.json",encoding="utf-8")#13
            file_sede = open( str(ruta)+"cyd_sede.json",encoding="utf-8")#14
            file = open(str(ruta)+"errores_finales2.txt", "w")
            data_file_escuela = json.load(file_escuela)#1
            data_file_asesoria = json.load(file_asesoria)#2
            data_file_asignacion = json.load(file_asignacion)#3
            data_file_calendario = json.load(file_calendario)#4
            data_file_crasistencia  = json.load(file_crasistencia)#5
            data_file_crhito = json.load(file_crhito)#6
            data_file_grupo = json.load(file_grupo)#7
            data_file_notaasistencia = json.load(file_notaasistencia)#8
            data_file_notahito = json.load(file_notahito)#9
            data_file_parescolaridad = json.load(file_parescolaridad)#10
            data_file_paretnia = json.load(file_paretnia)#11
            data_file_parrol = json.load( file_parrol)#12
            data_file_pargenero = json.load(file_pargenero)#13
            data_file_participante = json.load(file_participante)#14
            data_file_sede = json.load(file_sede)#15
            data_file_curso = json.load(file_curso)#16
            #data_ = json.load(f)
            """for data_escuela in data_file_escuela:
                 print("escuela:",str(data_escuela['id']))
                 try:
                      escuela= Escuela.objects.get(codigo=data_escuela["codigo"])
                      #print(escuela)
                 except Exception as a:
                      escuela = Escuela(
                           codigo=data_escuela["codigo"],
                           distrito=data_escuela["distrito"],
                           municipio=Municipio.objects.get(id=data_escuela["municipio_id"]),
                           nombre=data_escuela["nombre"],
                           direccion=data_escuela["direccion"],
                           telefono=data_escuela["telefono"],
                           nivel=EscNivel.objects.get(id=data_escuela["nivel_id"]),
                           sector=EscSector.objects.get(id=data_escuela["sector_id"]),
                           area=EscArea.objects.get(id=data_escuela["area_id"]),
                           status=EscStatus.objects.get(id=data_escuela["status_id"]),
                           modalidad=EscModalidad.objects.get(id=data_escuela["modalidad_id"]),
                           jornada=EscJornada.objects.get(id=data_escuela["jornada_id"]),
                           plan=EscPlan.objects.get(id=data_escuela["plan_id"]),                         
                           
                           )
                      escuela.save()
                      
                      file.write(str(a)+str(data_escuela["codigo"])+'\n')
                      #print(a)
            file_escuela.close()
            print("---FIN ESCUELA---")
            for data_parrol  in data_file_parrol:
                 print("rol:",data_parrol["id"])
                 try:
                      rol = cyd_m.ParRol.objects.get(id=data_parrol["id"])
                 except Exception as b:
                      file.write(str(b)+str(data_parrol["id"])+'\n')
                      rol = cyd_m.ParRol(
                           id = data_parrol["id"],
		                 nombre = data_parrol["nombre"]
                           
ZeroDivisionErrorError = models import                    )
                      rol.save()
            file_parrol.close()
            print("---FIN ROL---") 
            for data_participante  in data_file_participante:
                 print("participante",data_participante["id"])
                 try:
                      cyd_m.Participante.objects.id(id=data_participante["id"])
                 except Exception as c:
                    file.write(str(c)+str(data_participante["id"])+'\n')
                    nuevo_registro_participantes = cyd_m.Participante(
                            id = data_participante["id"],                            
                            nombre = data_participante["nombre"],                            
                            apellido= data_participante["apellido"],                            
                            genero = cyd_m.ParGenero.objects.get(id=data_participante["genero_id"]),                            
                            direccion = data_participante["direccion"],
                            mail = data_participante["mail"],                            
                            fecha_nac = data_participante["fecha_nac"],
                            escolaridad=cyd_m.ParEscolaridad.objects.get(id=data_participante["escolaridad_id"]),
                            escuela = Escuela.objects.get(id=data_participante["escuela_id"]),
                            etnia = cyd_m.ParEtnia.objects.get(id=data_participante["etnia_id"]),                            
                            rol = cyd_m.ParRol.objects.get(id=data_participante["rol_id"]),
                            slug= data_participante["slug"],
                            activo= data_participante["activo"],
                            dpi = data_participante["dpi"],  
                            tel_casa = data_participante["tel_casa"],
                            tel_movil = data_participante["tel_movil"]
                            
                              
                    )
                    nuevo_registro_participantes.save()
            file_participante.close()
            print("---FIN PARTICIPANTE---")
            for data_sede  in data_file_sede:
                 print("sede:",data_sede["id"])
                 try:
                     cyd_m.Sede.objects.get(id=data_sede["id"])
                 except Exception as d:
                      #print(datos["id_usr"])    
                      file.write(str(d)+str(data_sede["id"])+'\n')                  
                      sede=cyd_m.Sede(
                        id = data_sede["id"],
                        nombre = data_sede["nombre"],
                        capacitador = User.objects.get(id=data_sede["capacitador_id"]),
                        municipio = Municipio.objects.get(id=data_sede["municipio_id"]),
                        direccion = data_sede["direccion"],
                        observacion = data_sede["observacion"],
                        escuela_beneficiada = Escuela.objects.get(id=data_sede["escuela_beneficiada_id"]),
                        fecha_creacion  = data_sede["fecha_creacion"],
                        tipo_sede = data_sede["tipo_sede"],
                        activa = data_sede["activa"]                   
                        )
                      sede.save()
            file_sede.close()
            print("-- FIN SEDE----") """
            """for data_curso  in data_file_curso:
                 print("curso",data_curso["id"])
                 try:
                      curso = cyd_m.Curso.objects.get(id=data_curso["id"])
                 except Exception as f:
                      file.write(str(f)+str(data_curso["id"])+'\n')
                      curso = cyd_m.Curso(
                           id=data_curso["id"],
                           nombre=data_curso["nombre"],
                           nota_aprobacion=data_curso["nota_aprobacion"],                           
                           activo=data_curso["activo"]
                      )
                      curso.save()
            file_curso.close() 
            print("---FIN CURSO---")
            for data_grupo  in data_file_grupo:
                 print("grupo:",data_grupo["id"])
                 try:
                       cyd_m.Grupo.objects.get(id=data_grupo["id"])
                 except Exception as e:
                      file.write(str(e)+str(data_grupo["id"])+'\n')
                      grupo=cyd_m.Grupo(
                            id = data_grupo["id"],
                            sede = cyd_m.Sede.objects.get(id=data_grupo["sede_id"]),
                            numero =data_grupo["numero"] ,
                            curso = cyd_m.Curso.objects.get(id=data_grupo["curso_id"]),
                            comentario = data_grupo["comentario"],
                            cyd_grupo_creado_por = User.objects.get(id=data_grupo["cyd_grupo_creado_por_id"]),
                            activo = data_grupo["activo"]                     
                        )
                      grupo.save()
                    
            file_grupo.close()
            print("---FIN GRUPO---")"""
                 
            """for data_asignacion  in data_file_asignacion:
                 print("asignacion",data_asignacion["id"])
                 try:
                      cyd_m.Asignacion.objects.get(id=data_asignacion["id"])
                 except Exception as g:
                    asignacion=cyd_m.Asignacion(
                        id = data_asignacion["id"],
                        participante = cyd_m.Participante.objects.get(id=data_asignacion["participante_id"]),
                        grupo =cyd_m.Grupo.objects.get(id=data_asignacion["grupo_id"]),
                        abandono = data_asignacion["abandono"]                  
                    )
                    asignacion.save()
            file_asignacion.close()
            print("---FIN ASIGNACION---") """
            ### COMENTAR DESPUES DE SUBIR LAS NOTAS ####
            for data_crasistencia  in data_file_crasistencia:
                 print("crasitencia:",data_crasistencia["id"])
                 try:
                     cr_asi = cyd_m.CrAsistencia.objects.get(id=data_crasistencia["id"])
                 except Exception as h:
                      file.write(str(h)+str(data_crasistencia["id"])+'\n')
                      cr_asistecias=cyd_m.CrAsistencia(
                        id = data_crasistencia["id"],
                        curso=cyd_m.Curso.objects.get(id=data_crasistencia["curso_id"]),
                        modulo_num=data_crasistencia["modulo_num"],
                        punteo_max = data_crasistencia["punteo_max"]

                         )
                      cr_asistecias.save()            
            file_crasistencia.close()
            print("---FIN CRASISTENCIA---") 

            for data_crhito  in data_file_crhito:
                 print("cr_hito:",data_crhito["id"])
                 try:
                      cyd_m.CrHito.objects.get(id=data_crhito["id"])
                   
                 except Exception as i:
                     file.write(str(i)+str(data_crhito["id"])+'\n')
                     cr_hito=cyd_m.CrHito(
                        id = data_crhito["id"],
                        curso=cyd_m.Curso.objects.get(id=data_crhito["curso_id"]),
                        nombre=data_crhito["nombre"],
                        punteo_max=data_crhito["punteo_max"]
                     )
                     cr_hito.save()          
                 
            file_crhito.close()
            print("---FIN CRHITO---") 

            for data_calendario  in data_file_calendario:
                 print("calendario:",data_calendario["id"])
                 try:
                      cyd_m.Calendario.objects.get(id=data_calendario["id"])
                 except Exception as j:
                      file.write(str(j)+str(data_calendario["id"])+'\n')
                      calendario=cyd_m.Calendario(
                            id = data_calendario["id"],
                            cr_asistencia=cyd_m.CrAsistencia.objects.get(id=data_calendario["cr_asistencia_id"]),
                            grupo=cyd_m.Grupo.objects.get(id=data_calendario["grupo_id"]),
                            fecha=data_calendario["fecha"],
                            hora_inicio=data_calendario["hora_inicio"],
                            hora_fin=data_calendario["hora_fin"]

                            )
                      calendario.save()
                      
            file_calendario.close()
            print("---FIN CALENDARIO---")

            ##### HASTA ACA ####

            """for data_notaasistencia  in data_file_notaasistencia:
                 print("nota asistencia:",data_notaasistencia["id"])
                 try:
                     asistencia=  cyd_m.NotaAsistencia.objects.get(id=data_notaasistencia ["id"])                     
                 except Exception as k:
                    file.write(str(k)+str(data_notaasistencia["id"])+'\n')
                    notas_asistencias=cyd_m.NotaAsistencia(
                            id = data_notaasistencia ["id"],
                            asignacion=cyd_m.Asignacion.objects.get(id=data_notaasistencia ["asignacion_id"]),
                            gr_calendario=cyd_m.Calendario.objects.get(id=data_notaasistencia ["gr_calendario_id"]),
                            nota=data_notaasistencia ["nota"]

                         )
                    notas_asistencias.save()
                         
            file_notaasistencia.close()
            print("---FIN NOTA ASISTENCIA---")   """   
            """for data_notahito  in data_file_notahito:
                 print("notahito",data_notahito["id"])
                 
                 try:
                       nota = cyd_m.NotaHito.objects.get(id=data_notahito["id"])                       
                 except Exception as l:
                        file.write(str(l)+str(data_notahito["id"])+'\n')
                        nota_hito=cyd_m.NotaHito(
                            id = data_notahito["id"],
                            asignacion=cyd_m.Asignacion.objects.get(id=data_notahito["asignacion_id"]),
                            cr_hito = cyd_m.CrHito.objects.get(id=data_notahito["cr_hito_id"]),
                            nota=data_notahito["nota"]

                        )
                        nota_hito.save()
            file_notahito.close()
            print("---FIN HITO---")"""
          ### DESCOMENTAR ESTO CARLITOS DESPUES DE TERMINAR DE SUBIR LAS NOTAS 
            

            """for data_asesoria  in data_file_asesoria:
                 print("asesoria:",data_asesoria["id"])
                 try:
                      aseso = cyd_m.Asesoria.objects.get(id=data_asesoria["id"])
                 except Exception as m:
                    file.write(str(m)+str(data_asesoria["id"])+'\n')
                    asesorias=cyd_m.Asesoria(
                        id = data_asesoria["id"],
                        sede=cyd_m.Sede.objects.get(id=data_asesoria["sede_id"]),
                        fecha=data_asesoria["fecha"],
                        hora_inicio=data_asesoria["hora_inicio"],
                        hora_fin=data_asesoria["hora_fin"],
                        observacion=data_asesoria["observacion"]

                    )
                    asesorias.save()
            file_asesoria.close()
            print("---FIN ASESORIA---") """

            #####HASTA  ACA ################################    
                
                
                       
           
                
                  
            """ for data_parescolaridad  in data_file_parescolaridad:
                 print(data_parescolaridad["id"])
            file_parescolaridad.close()
            print("---FIN ESCOLARIDAD---")      
            for data_paretnia  in data_file_paretnia:
                 print(data_paretnia["id"])
            file_paretnia.close()
            print("--FIN ESCOLARIDAD----")      
            for data_pargenero  in data_file_pargenero:
                 print(data_pargenero["id"])
            file_pargenero.close()
            print("--PAR GENERO----")      
             """    
                  
                 
            file.close()
                            
            return Response(
                            "Datos ingresados Correctamente Carlitos Ya podes ir a descanzar",
                            status=status.HTTP_200_OK
                        ) 
class RevionErrores(views.APIView):
        """ Importar registros de excel con DjangoRestFramework para Bienestar
        """
        def get(self, request):
            ruta = "C:/Users/Edgar/Documents/TablasMigrarSuni/json/" 
            file = open(str(ruta)+"errores_finales2.txt", "rb")
            lineas = file.readlines()
                               
            return Response(
                            lineas,
                            status=status.HTTP_200_OK
                        )

class SubirAsignacionesJson(views.APIView):
        """ Importar registros de un json de `cyd_Asignaciones`del suni de pruebas  para la tabla `cyd_m.Asignaciones` del SUNI de produccion
        """
        def get(self, request): 
            ruta = "C:/Users/Edgar/Documents/TablasMigrarSuni/json/Asignaciones/"
            nombre_archivo = str("cyd_asignaciones") +str(self.request.GET['archivo'])+".json"
            ruta_archivo = ruta + nombre_archivo            
            if path.exists(ruta_archivo):
                 file = open(ruta_archivo,encoding="utf-8")
                 data = json.load(file)
                 for data_asignacion  in data:
                    #print("asignacion",data_asignacion["id"])
                    try:
                         asignacion = cyd_m.Asignacion.objects.get(id=data_asignacion["id"])
                         if asignacion:
                              mensaje = "El archivo: {} YA FUE CARGADO CON EXITO por favor ingrese otro numero de archivo".format(nombre_archivo)
                              return Response(
                                        mensaje,
                                        status=status.HTTP_404_NOT_FOUND
                                   )
                    except Exception as g:
                         asignacion=cyd_m.Asignacion(
                         id = data_asignacion["id"],
                         participante = cyd_m.Participante.objects.get(id=data_asignacion["participante_id"]),
                         grupo =cyd_m.Grupo.objects.get(id=data_asignacion["grupo_id"]),
                         abandono = data_asignacion["abandono"]                  
                         )
                         asignacion.save()
                 
                 mensaje = "Datos del arcnivo: {} ingresados Correctamente pase al siguiente archivo si existe".format(nombre_archivo)
                 return Response(
                            mensaje,
                            status=status.HTTP_200_OK
                        )
            else:
                 mensaje = "El archivo: {} no existe por favor ingrese otro numero de archivo".format(nombre_archivo)
                 return Response(
                            mensaje,
                            status=status.HTTP_404_NOT_FOUND
                        )
            
class SubirNotasAsistenciaJson(views.APIView):
        """ Importar registros de un json de `cyd_NotaAsistencia`del suni de pruebas  para la tabla `cyd_m.NotaAsistencia` del SUNI de produccion
        """
        def get(self, request): 
            ruta = "C:/Users/Edgar/Documents/TablasMigrarSuni/json/NotasAsistencia/"
            nombre_archivo = str("cyd_notaasistencia") +str(self.request.GET['archivo'])+".json"
            ruta_archivo = ruta + nombre_archivo           
            if path.exists(ruta_archivo):                
                 file = open(ruta_archivo,encoding="utf-8")
                 data_file = json.load(file)
                 for data_notaasistencia  in data_file:
                    #print("nota asistencia:",data_notaasistencia["id"])
                    try:
                         asistencia=  cyd_m.NotaAsistencia.objects.get(id=data_notaasistencia ["id"])
                         if asistencia:
                              mensaje = "El archivo: {} YA FUE CARGADO CON EXITO por favor ingrese otro numero de archivo".format(nombre_archivo)
                              return Response(
                                        mensaje,
                                        status=status.HTTP_404_NOT_FOUND
                                   )                     
                    except Exception as k:
                         notas_asistencias=cyd_m.NotaAsistencia(
                              id = data_notaasistencia ["id"],
                              asignacion=cyd_m.Asignacion.objects.get(id=data_notaasistencia ["asignacion_id"]),
                              gr_calendario=cyd_m.Calendario.objects.get(id=data_notaasistencia ["gr_calendario_id"]),
                              nota=data_notaasistencia ["nota"]

                              )
                         notas_asistencias.save()
                 
                 mensaje = "Datos del arcnivo: {} ingresados Correctamente pase al siguiente archivo si existe".format(nombre_archivo)
                 return Response(
                            mensaje,
                            status=status.HTTP_200_OK
                        )
            else:
                 mensaje = "El archivo: {} no existe por favor ingrese otro numero de archivo".format(nombre_archivo)
                 return Response(
                            mensaje,
                            status=status.HTTP_404_NOT_FOUND
                        )
            

class SubirNotasHitosJson(views.APIView):
        """ Importar registros de un json de `cyd_m.NotasHito`del suni de pruebas  para la tabla `cyd_m.NotasHitos` al SUNI de produccion
        """
        def get(self, request): 
            ruta = "C:/Users/Edgar/Documents/TablasMigrarSuni/json/NotasHitos/"
            nombre_archivo = str("cyd_notahito") +str(self.request.GET['archivo'])+".json"
            ruta_archivo = ruta + nombre_archivo           
            if path.exists(ruta_archivo):                
                 file = open(ruta_archivo,encoding="utf-8")
                 data_file = json.load(file)
                 for data_notahito  in data_file:
                    #print("notahito",data_notahito["id"])                    
                    try:
                         nota = cyd_m.NotaHito.objects.get(id=data_notahito["id"])
                         if nota:
                              mensaje = "El archivo: {} YA FUE CARGADO CON EXITO por favor ingrese otro numero de archivo".format(nombre_archivo)
                              return Response(
                                        mensaje,
                                        status=status.HTTP_404_NOT_FOUND
                                   )                       
                    except Exception as l:
                         nota_hito=cyd_m.NotaHito(
                              id = data_notahito["id"],
                              asignacion=cyd_m.Asignacion.objects.get(id=data_notahito["asignacion_id"]),
                              cr_hito = cyd_m.CrHito.objects.get(id=data_notahito["cr_hito_id"]),
                              nota=data_notahito["nota"]

                         )
                         nota_hito.save()
                 mensaje = "Datos del arcnivo: {} ingresados Correctamente pase al siguiente archivo si existe".format(nombre_archivo)
                 return Response(
                            mensaje,
                            status=status.HTTP_200_OK
                        )
            else:
                 mensaje = "El archivo: {} no existe por favor ingrese otro numero de archivo".format(nombre_archivo)
                 return Response(
                            mensaje,
                            status=status.HTTP_404_NOT_FOUND
                        )


#Fin primera migración 

#Inicio de segunda migración 
class escuelasApi(views.APIView):
    """Comentario de Api de prueba para la Apis de django"""
    def get(self, request): 
     archivo_excel_path = "C:/Users/PC/Desktop/Migracion2/Escuelas_MIG_2.xlsx"         #Aquí va la dirección
     archivo_excel = load_workbook(filename=archivo_excel_path)
     hoja_excel = archivo_excel.active
     max_row = hoja_excel.max_row

     existentes = 0
     agregadas = 0 
     ruta_archivo_txt = os.path.splitext(archivo_excel_path)[0] + "_log.txt"

     with open(ruta_archivo_txt, "w") as archivo_log:
          for i in range(2, max_row + 1):
               try:
                    escuela = cyd_m.Escuela.objects.get(codigo=hoja_excel.cell(row=i, column=2).value)
                    existentes += 1
                    print(str(i) + " -> " + str(escuela))

               except cyd_m.Escuela.DoesNotExist:
                    nueva_escuela = cyd_m.Escuela(
                        codigo=hoja_excel.cell(row=i, column=2).value,
                        distrito=hoja_excel.cell(row=i, column=3).value,
                        municipio=esc_m.Municipio.objects.get(id=hoja_excel.cell(row=i, column=10).value), 
                        nombre=hoja_excel.cell(row=i, column=4).value,
                        direccion=hoja_excel.cell(row=i, column=5).value,
                        telefono=hoja_excel.cell(row=i, column=6).value,
                        nivel=esc_m.EscNivel.objects.get(id=hoja_excel.cell(row=i, column=11).value),
                        sector=esc_m.EscSector.objects.get(id=hoja_excel.cell(row=i, column=13).value),
                        area=esc_m.EscArea.objects.get(id=hoja_excel.cell(row=i, column=7).value),
                        status=esc_m.EscStatus.objects.get(id=hoja_excel.cell(row=i, column=14).value),
                        modalidad=esc_m.EscModalidad.objects.get(id=hoja_excel.cell(row=i, column=9).value),
                        jornada=esc_m.EscJornada.objects.get(id=hoja_excel.cell(row=i, column=8).value),
                        plan=esc_m.EscPlan.objects.get(id=hoja_excel.cell(row=i, column=12).value),
                        esc_creado_por=User.objects.get(id=hoja_excel.cell(row=i, column=16).value)
                    )
                    nueva_escuela.save()
                    agregadas += 1
                    mensaje = "{} Escuela creada: {}".format(i, hoja_excel.cell(row=i, column=4).value)
                    print(mensaje)
                    archivo_log.write(mensaje + "\n")

          resumen = "Total creadas: {}, Ya existentes: {}".format(agregadas, existentes)
          print(resumen)
          archivo_log.write(resumen + "\n")
          print("---FIN ESCUELAS---")

        
          return Response(
               "Datos ingresados con exito", 
               status=status.HTTP_200_OK
          )
    


class equipamientoApi(views.APIView):
    """Comentario de Api de prueba para la Apis de django"""
    def get(self, request): 
          archivo_excel_path = "C:/Users/PC/Desktop/Migracion2/Equipamientos_MIG_2.xlsx"
          archivo_excel = load_workbook(filename=archivo_excel_path)
          hoja_excel = archivo_excel.active
          max_row = hoja_excel.max_row
          min_row = hoja_excel.min_row

          equipadas = 0
          no_equipadas = 0
          no_existe = 0 
          ruta_archivo_txt = os.path.splitext(archivo_excel_path)[0] + "_log.txt"
          j = 1 

          with open(ruta_archivo_txt, "w") as archivo_log:
               for i in range(2, max_row + 1):
                    try:
                         escuela = cyd_m.Escuela.objects.get(codigo = hoja_excel.cell(row = i, column = 2).value)
                         

                         if escuela.equipada:
                              equipadas += 1

                         else:
                              ultimo_id = tpe_m.Equipamiento.objects.latest('id').id
                              ultimo_referencia = tpe_m.Equipamiento.objects.filter( cooperante = 172).last()  #Cooperante BEQT
                                  
                              nuevo_equipamiento = tpe_m.Equipamiento.objects.create(
                                   id = ultimo_id + 1,
                                   no_referencia = int(ultimo_referencia.no_referencia) + j, 
                                   estado = tpe_m.EquipamientoEstado.objects.get(id = "1"), 
                                   escuela = cyd_m.Escuela.objects.get(codigo = hoja_excel.cell(row = i, column = 2).value), 
                                   fecha = "2010-01-01", 
                                   observacion = "Historico de equipamientos de 2010", 
                                   renovacion = False, 
                                   servidor_khan = False,
                                   creado_por = User.objects.get(id = hoja_excel.cell(row = i, column = 4).value),
                                   #cooperante = Cooperante.objects.get(nombre = "FUNSEPA")
                              )
     
                              registro = "{} Escuela equipada: {}".format(i, hoja_excel.cell(row=i, column=3).value)
                              archivo_log.write(registro + "\n")
                              nuevo_equipamiento.save()
                              no_equipadas += 1
                              print(str(i) + " Creada: " +  hoja_excel.cell(row = i, column = 2).value)
                              j += 1

                    except Exception as c:
                         print(c)
                         no_existe += 1

               resumen = "Total existentes: {}, total creadas: {}, no econtradas {}".format(equipadas, no_equipadas, no_existe)
               archivo_log.write(resumen + "\n")
               print(resumen)
               print("---FIN EQUIPAMIENTOS---")
        
          return Response(
               "Datos ingresados con exito", 
               status=status.HTTP_200_OK
          )


class participantesApi(views.APIView):
    """Comentario de Api de prueba para la Apis de django class:`Participante`"""

    def get(self, request):
          archivo = self.request.GET.get('archivo')
          nombre_archivo = "C:/Users/dguty/Desktop/Migracion2/Participantes_MIG_2A_Final" + archivo.upper() + ".xlsx" 
          archivo_excel_path = filename= nombre_archivo
          archivo_excel = load_workbook(filename=archivo_excel_path)
          hoja_excel = archivo_excel.active
          max_row = hoja_excel.max_row
          min_row = hoja_excel.min_row

          existentes = 0
          agregados = 0 
          ruta_archivo_txt = os.path.splitext(archivo_excel_path)[0] + "_log.txt"

          with open(ruta_archivo_txt, "w") as archivo_log:
               for i in range(2, max_row + 1):
                    try:
                         if cyd_m.Participante.objects.get(nombre = hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value) : 
                              participante = cyd_m.Participante.objects.get(nombre = hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value)
                              registro = "{} Existe -> {} {}".format(i, hoja_excel.cell(row=i, column=4).value, hoja_excel.cell(row=i, column=5).value)
                              archivo_log.write(registro + "\n")
                              existentes +=1
                    except Exception as c:
                         registro_participantes = cyd_m.Participante(
                         dpi = hoja_excel.cell(row = i, column = 3).value,                          
                         nombre = hoja_excel.cell(row = i, column = 4).value,                            
                         apellido= hoja_excel.cell(row = i, column = 5).value,                          
                         genero = cyd_m.ParGenero.objects.get(id= hoja_excel.cell(row = i, column = 7).value), 
                         rol = cyd_m.ParRol.objects.get(id = hoja_excel.cell(row = i, column = 6).value), 
                         escuela = Escuela.objects.get(codigo = hoja_excel.cell(row = i, column = 2).value),  
                         direccion = hoja_excel.cell(row = i, column = 8).value,  
                         mail = hoja_excel.cell(row = i, column = 9).value,
                         tel_casa = hoja_excel.cell(row = i, column = 10).value,  
                         tel_movil = hoja_excel.cell(row = i, column = 11).value,                          
                         fecha_nac = hoja_excel.cell(row = i, column = 12).value,
                         )
                         registro_participantes.save()
                         agregados +=1
                         print("Participante creado: " + str(i) + " " + hoja_excel.cell(row = i, column = 4).value)
                         
               resumen = "Total existentes: {}, total creadas: {}".format(existentes, agregados)
               print(resumen)
               archivo_log.write(resumen + "\n")
          
          return Response(
               "Datos ingresados con exito", 
               status=status.HTTP_200_OK
          )
    
class SedeApi(views.APIView): 
    """Comentario de Api de prueba para la Apis de django"""

    def get(self, request): 
          archivo_excel_path = filename="C:/Users/dguty/Desktop/Migracion2/Sedes_MIG_2.xlsx"  #Aquí va la dirección 
          archivo_excel = load_workbook(filename=archivo_excel_path)
          hoja_excel = archivo_excel.active
          max_row = hoja_excel.max_row

          existentes = 0
          agregados = 0 
          ruta_archivo_txt = os.path.splitext(archivo_excel_path)[0] + "_log.txt"

          with open(ruta_archivo_txt, "w") as archivo_log:
               for i in range(2, max_row + 1):
                    try:
                         if cyd_m.Sede.objects.get(escuela_beneficiada__codigo = hoja_excel.cell(row = i, column = 8).value, capacitador__id = hoja_excel.cell(row = i, column = 2).value):
                              existentes +=1

                    except ObjectDoesNotExist:
                         try:
                              escuela = esc_m.Escuela.objects.get(codigo=hoja_excel.cell(row=i, column=8).value)

                              registro_sede = cyd_m.Sede(
                                   nombre=escuela.nombre,
                                   capacitador=User.objects.get(id=hoja_excel.cell(row=i, column=2).value),
                                   municipio=escuela.municipio,
                                   direccion=escuela.direccion,
                                   observacion=hoja_excel.cell(row=i, column=5).value,
                                   mapa=escuela.mapa,
                                   escuela_beneficiada=escuela,
                                   tipo_sede=hoja_excel.cell(row=i, column=9).value,
                                   fecha_creacion=hoja_excel.cell(row=i, column=10).value,

                              )
                              registro_sede.save()
                              agregados += 1
                              print("Sede creada: " + str(i) + " " + hoja_excel.cell(row=i, column=8).value)
                              registro = "{} Sede creada: {}".format(i, hoja_excel.cell(row=i, column=8).value)
                              archivo_log.write(registro + "\n")
                         except Exception as e:
                              print("No se pudo crear la sede para la escuela con el código :" +  hoja_excel.cell(row=i, column=8).value)

               resumen = "Total existentes: {}, total creadas: {}".format(existentes, agregados)
               print(resumen)
               archivo_log.write(resumen + "\n")

          return Response(
               "Datos ingresados con exito", 
               status=status.HTTP_200_OK
          )
    

class GrupoApi(views.APIView): 
    """Comentario de Api de prueba para la Apis de django"""

    def get(self, request): 
          archivo_excel_path = filename= "C:/Users/dguty/Desktop/Migracion2/Grupo_MIG_2.xlsx" #Aquí va la dirección 
          archivo_excel = load_workbook(filename=archivo_excel_path)
          hoja_excel = archivo_excel.active
          max_row = hoja_excel.max_row

          existentes = 0
          agregados = 0 
          ruta_archivo_txt = os.path.splitext(archivo_excel_path)[0] + "_log.txt"

          with open(ruta_archivo_txt, "w") as archivo_log:
               for i in range(2, max_row + 1):
                    try:
                         grupo = cyd_m.Grupo.objects.get(sede__escuela_beneficiada__codigo = hoja_excel.cell(row = i, column = 1).value, curso__id = hoja_excel.cell(row = i, column = 7).value, sede__capacitador__id = hoja_excel.cell(row = i, column = 6).value)
                         existentes +=1

                    except ObjectDoesNotExist:
                         try: 
                              sede_mas_antigua = cyd_m.Sede.objects.filter(escuela_beneficiada__codigo=hoja_excel.cell(row=i, column=1).value, capacitador_id = hoja_excel.cell(row = i, column = 6).value).order_by('fecha_creacion').first()
                              registro_grupo = cyd_m.Grupo(
                                   sede = sede_mas_antigua,
                                   numero = hoja_excel.cell(row = i, column = 2).value,
                                   curso = cyd_m.Curso.objects.get(id = hoja_excel.cell(row = i, column = 7).value),
                                   comentario = "Grupo creado con información tomada de las Actas de promoción y bases de datos del histórico de Capacitación y Proyectos. Migrado al SUNI en el 2024",
                                   activo = True
                              )
                              registro_grupo.save()
                              agregados +=1
                              print(str(i) + " Grupo creado: " + hoja_excel.cell(row = i, column = 1).value)
                              registro = "{} Sede creada: {}".format(i, hoja_excel.cell(row=i, column=1).value)
                              archivo_log.write(registro + "\n")

                         except Exception as c:
                              print("No se pudo crear el grupo: " +  str(c))

               resumen = "Total existentes: {}, total creadas: {}".format(existentes, agregados)
               archivo_log.write(resumen + "\n")
               print(resumen)
        
          return Response(
               "Datos ingresados con exito", 
               status=status.HTTP_200_OK
          )
    

class AsignacionesApi(views.APIView): 
    """Comentario de Api de prueba para la Apis de django class ´Asignacion´"""
    def get(self, request): 
          archivo = self.request.GET.get('archivo')
          nombre_archivo = "C:/Users/dguty/Desktop/Migracion2/Asignacion_MIG_2A_Final" + archivo.upper() + ".xlsx" 
          archivo_excel_path = filename= nombre_archivo
          print(archivo_excel_path)
          archivo_excel = load_workbook(filename=archivo_excel_path)
          hoja_excel = archivo_excel.active
          max_row = hoja_excel.max_row

          existentes = 0
          agregados = 0 
          ruta_archivo_txt = os.path.splitext(archivo_excel_path)[0] + "_log.txt"

          with open(ruta_archivo_txt, "w") as archivo_log:
               for i in range(2, max_row + 1):
                    try:
                         escuela_par = cyd_m.Escuela.objects.get(codigo = hoja_excel.cell(row = i, column = 2).value)

                         try:
                              asignado = cyd_m.Participante.objects.get(nombre =  hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value, escuela = escuela_par)

                         except ObjectDoesNotExist: 
                              asignado = cyd_m.Participante.objects.get(nombre = hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value)

                         except MultipleObjectsReturned: 
                              asignado = cyd_m.Participante.objects.filter(nombre =  hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value, escuela = escuela_par).last()

                         cyd_m.Asignacion.objects.get(participante = asignado , grupo__curso__id= hoja_excel.cell(row = i, column = 28, grupo__sede__capacitador__id = hoja_excel.cell(row=i, column=30).value).value)
                         print(cyd_m.Asignacion.objects.get(participante = asignado , grupo__curso__id= hoja_excel.cell(row = i, column = 28).value))
                         existentes +=1

                    except MultipleObjectsReturned: 
                         print(str(i) + "Participante " + str(asignado) + " asignado antes")
                         existentes +=1

                    except Exception as c:
                         escuela_par = cyd_m.Escuela.objects.get(codigo = hoja_excel.cell(row = i, column = 2).value)

                         try:
                              par_asignado = cyd_m.Participante.objects.get(nombre =  hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value, escuela = escuela_par)
                         except ObjectDoesNotExist: 
                              par_asignado = cyd_m.Participante.objects.get(nombre = hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value)
                         except MultipleObjectsReturned: 
                              par_asignado = cyd_m.Participante.objects.filter(nombre =  hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value, escuela = escuela_par).last()
                         
                         sede_mas_antigua = cyd_m.Sede.objects.filter(escuela_beneficiada__codigo=hoja_excel.cell(row=i, column=27).value, capacitador = hoja_excel.cell(row=i, column=30).value).order_by('fecha_creacion').last()
                         grupo = cyd_m.Grupo.objects.get(curso__id = int(hoja_excel.cell(row = i, column = 28).value) , sede = sede_mas_antigua, sede__capacitador__id = hoja_excel.cell(row=i, column=30).value)
                         
                         registro_asigna = cyd_m.Asignacion(
                              participante = par_asignado,
                              grupo = grupo,
                              abandono = False
                         )
                         registro_asigna.save()
                         agregados +=1
                         print("Asiganción creada: " + str(i) + " " + hoja_excel.cell(row = i, column = 2).value)
                         registro = "{} Sede creada: {}".format(i, hoja_excel.cell(row=i, column=2).value)
                         archivo_log.write(registro + "\n")

               resumen = "Total existentes: {}, total creadas: {}".format(existentes, agregados)
               archivo_log.write(resumen + "\n")
               print(resumen)
        
          return Response(
               "Datos ingresados con exito", 
               status=status.HTTP_200_OK
          )
    
class HitosApi(views.APIView): 
    """Comentario de Api de prueba para la Apis de django :class:`NotaHito`"""

    def get(self, request): 
          archivo = self.request.GET.get('archivo')
          nombre_archivo = "C:/Users/dguty/Desktop/Migracion2/Asignacion_MIG_2A_Final" + archivo.upper() + ".xlsx" 
          print(nombre_archivo)
          archivo_excel_path = filename= nombre_archivo
          archivo_excel = load_workbook(filename=archivo_excel_path)
          hoja_excel = archivo_excel.active
          max_row = hoja_excel.max_row

          existentes = 0
          agregados = 0 
          problematicos = 0
          ruta_archivo_txt = os.path.splitext(archivo_excel_path)[0] + "_log.txt"

          with open(ruta_archivo_txt, "w") as archivo_log:
               for i in range(2, max_row + 1):
                    try:
                         
                         escuela_par = cyd_m.Escuela.objects.get(codigo = hoja_excel.cell(row = i, column = 2).value)
                         try:
                              asignado = cyd_m.Participante.objects.get(nombre =  hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value, escuela = escuela_par)
                         except ObjectDoesNotExist: 
                              asignado = cyd_m.Participante.objects.get(nombre = hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value)
                         except MultipleObjectsReturned: 
                              asignado = cyd_m.Participante.objects.filter(nombre =  hoja_excel.cell(row = i, column = 4).value, apellido = hoja_excel.cell(row = i, column = 5).value, escuela = escuela_par).last()
                         
                         sede_mas_antigua = cyd_m.Sede.objects.filter(escuela_beneficiada__codigo=hoja_excel.cell(row=i, column=27).value, capacitador = hoja_excel.cell(row=i, column=30).value).order_by('fecha_creacion').first()
                         grupo = cyd_m.Grupo.objects.get(curso__id = hoja_excel.cell(row = i, column = 28).value , sede = sede_mas_antigua, sede__capacitador = hoja_excel.cell(row=i, column=30).value )
                         asignacion = cyd_m.Asignacion.objects.get(participante = asignado , grupo = grupo)

                         nota = cyd_m.NotaHito.objects.get(asignacion = asignacion)
                         nota.nota = hoja_excel.cell(row = i, column = 23).value
                         nota.save()
                         print(i)
                         agregados +=1

                    except Exception as c:
                         print(c)
                         problematicos += 1


               resumen = "Total existentes: {}, total creadas: {}".format(existentes, agregados)
               archivo_log.write(resumen + "\n")
               print(resumen)

          return Response(
               "Datos ingresados con exito", 
               status=status.HTTP_200_OK
          )
    
#Fin apis de segunda parte migración Historico 2010

###############################################
class CapacitacionNotas(views.APIView):
        """ Importar registros de excel con DjangoRestFramework para Capacitacion
        """
        
        def get(self, request):          
            wb_obj = load_workbook(filename = "CSV_SUNI_OLD/notas_notas.xlsx")            
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            m_col= sheet_obj.max_column
            for i in range(2, m_row+1):
                id=sheet_obj.cell(row=i, column=1).value
                nota = sheet_obj.cell(row=i, column=6).value
                print(id,"->",nota)                            
            return Response(
                "Datos ingresados Correctamente",
                status=status.HTTP_200_OK
            )



###############################################
class estadoFormularioAPI(views.APIView):
    """Toma los datos del modulo de Evaluación para crear los modelos en la clase: ´estadoFormulario´ """

    def get(self, request):          
          asignaciones = AsignacionPregunta.objects.all()
          formularios = Formulario.objects.all()
          agrupacion = {}
          formularios_creados = 0
          contar = 0 


          for formulario in formularios:
               sede = cyd_m.Sede.objects.get(capacitador = formulario.usuario, escuela_beneficiada = formulario.escuela, activa = True)
               formulario.sede = sede
               formulario.save()

          for asignacion in asignaciones:
               participante = asignacion.evaluado

               if participante not in agrupacion:
                    agrupacion[participante] = []
               agrupacion[participante].append(asignacion)

          for participante, asignaciones in agrupacion.items():
               all_respondidas = all(asignacion.respondido for asignacion in asignaciones)

               for asignacion in asignaciones:
                    contar += 1

               estado_formulario = estadoFormulario.objects.create()
               estado_formulario.preguntas.set(asignaciones)
               estado_formulario.estado = all_respondidas

               estado_formulario.save()
               formularios_creados += 1
               
          print('Se crearon {} formularios'.format(formularios_creados))

          return Response("Objetos creados correctamente: {}".format(formularios_creados), status=status.HTTP_200_OK)

