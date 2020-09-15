from django.shortcuts import render
from django.views.generic import TemplateView, FormView, DetailView, View, CreateView
from django.middleware import csrf
from apps.Bienestar.models import Colaborador
from apps.Bienestar import forms as bienestar_f
from django.urls import reverse_lazy
from openpyxl import load_workbook
from django.http import HttpResponse, HttpResponseRedirect
from rest_framework import views, status
from rest_framework.response import Response
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.models import User
from braces.views import (
    LoginRequiredMixin)
from django.core import serializers
import dateutil.parser
import datetime
import dateutil.parser
import os
from django.conf import settings
from django.core.files.storage import FileSystemStorage

# Create your views here.
class BienestarExcelAddView(LoginRequiredMixin, TemplateView):
    template_name = 'bienestar/cargar_excel.html'
    def post(self,request):
        if request.method == 'POST' and request.FILES['myfile']:
            file_list =os.listdir(settings.MEDIA_ROOT_EXCEL_BIENESTAR)
            number_file=len(file_list)
            myfile=request.FILES['myfile']
            new_name=str('Bienestar') + str(number_file) +str(".")+str(myfile.name.split(".")[-1])
            print(new_name)
            fs=FileSystemStorage(location=settings.MEDIA_ROOT_EXCEL_BIENESTAR)
            filename=fs.save(new_name, myfile)
            uploaded_file_url=fs.url(filename)
            return render(request, 'bienestar/cargar_excel.html',{
                'uploaded_file_url':uploaded_file_url
        })
        return render(request,'bienestar/cargar_excel.html')

class CuestionarioView(LoginRequiredMixin, CreateView):
    """ Vista  encargada de obtener el dpi del participante por medio un  formulario y  muesta el listado de los
    cursos que se tienen asignados
    """
    template_name = 'bienestar/bienestar.html'
    form_class = bienestar_f.CuestionarioForm

    def get_success_url(self):
        return reverse_lazy('cuestionario')

class BienestarListView(LoginRequiredMixin, FormView):
    """Vista   para obtener los datos de PrecioEstandar mediante una :class:`PrecioEstandar`
    Funciona  para recibir los datos de un  'PrecioEstandarInformeForm' mediante el metodo  POST.  y
    nos muestra el template de visitas mediante el metodo GET.
    """
    model = Colaborador
    template_name = 'bienestar/bienestar_informe.html'
    form_class = bienestar_f.BienestarInformeForm

class ResultadoBienestarJson(LoginRequiredMixin, views.APIView):
        """ Importar registros de excel con DjangoRestFramework para Bienestar
        """
        def get(self, request):
            file_list =os.listdir(settings.MEDIA_ROOT_EXCEL_BIENESTAR)
            number_file=len(file_list)
            name_file =file_list[number_file-1]
            ruta = str(settings.MEDIA_ROOT_EXCEL_BIENESTAR)+str(name_file)
            listado_datos=[]
            wb_obj = load_workbook(filename = ruta)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            m_col= sheet_obj.max_column
            for i in range(2, m_row+1):
                marca_temporal=sheet_obj.cell(row=i, column=1).value
                correo=sheet_obj.cell(row=i, column=2).value
                if Colaborador.objects.filter(fecha=marca_temporal,email=correo).exists():
                    print("Si existe el colaborador")
                else:
                    print("El colaborador no existe")
                    nombre=sheet_obj.cell(row=i, column=3).value
                    edad=sheet_obj.cell(row=i, column=4).value
                    dpi=sheet_obj.cell(row=i, column=5).value
                    pregunta1=sheet_obj.cell(row=i, column=6).value
                    pregunta2=sheet_obj.cell(row=i, column=7).value
                    pregunta3=sheet_obj.cell(row=i, column=8).value
                    pregunta4=sheet_obj.cell(row=i, column=9).value
                    pregunta5=sheet_obj.cell(row=i, column=10).value
                    pregunta6=sheet_obj.cell(row=i, column=11).value
                    pregunta7=sheet_obj.cell(row=i, column=12).value
                    pregunta8=sheet_obj.cell(row=i, column=13).value
                    pregunta9=sheet_obj.cell(row=i, column=14).value
                    pregunta10=sheet_obj.cell(row=i, column=15).value
                    pregunta11=sheet_obj.cell(row=i, column=16).value
                    pregunta12=sheet_obj.cell(row=i, column=17).value
                    pregunta13=sheet_obj.cell(row=i, column=18).value
                    pregunta14=sheet_obj.cell(row=i, column=19).value
                    pregunta15=sheet_obj.cell(row=i, column=20).value
                    pregunta16=sheet_obj.cell(row=i, column=21).value
                    pregunta17=sheet_obj.cell(row=i, column=22).value
                    ingresar_colaborador=Colaborador(
                        fecha=marca_temporal,
                        email=correo,
                        usuario=nombre,
                        edad=edad,
                        dpi=dpi,
                        pregunta1=pregunta1,
                        pregunta2=pregunta2,
                        pregunta3=pregunta3,
                        pregunta4=pregunta4,
                        pregunta5=pregunta5,
                        pregunta6=pregunta6,
                        pregunta7=pregunta7,
                        pregunta8=pregunta8,
                        pregunta9=pregunta9,
                        pregunta10=pregunta10,
                        pregunta11=pregunta11,
                        pregunta12=pregunta12,
                        pregunta13=pregunta13,
                        pregunta14=pregunta14,
                        pregunta15=pregunta15,
                        pregunta16=pregunta16,
                        pregunta17=pregunta17
                    )
                    ingresar_colaborador.save()
            return Response(
                "Datos ingresados Correctamente",
                status=status.HTTP_200_OK
            )
class InformeBienestarJson(LoginRequiredMixin, views.APIView):
        """ Regreso los datos obtenidos del modelo de `Bienestar` para generar el informe
        """
        def post(self, request):
            datos_nuevos={}
            datos_enviar_json=[]
            fecha_min =self.request.POST['fecha_min']
            fecha_max = self.request.POST['fecha_max']
            nombre = self.request.POST['colaborador']
            correo=Colaborador.objects.get(usuario=nombre)
            datos_colaborador=Colaborador.objects.filter(email=correo.email, fecha__gte=fecha_min,fecha__lte=fecha_max).first()
            datos_buscar_colaborador=Colaborador.objects.filter(email=correo.email, fecha__gte=fecha_min,fecha__lte=fecha_max)
            respuesta_si=Colaborador.objects.filter(email=correo.email, fecha__gte=fecha_min,fecha__lte=fecha_max,pregunta4='Sí').count()
            respuesta_no=Colaborador.objects.filter(email=correo.email, fecha__gte=fecha_min,fecha__lte=fecha_max,pregunta4='No').count()            
            datos_nuevos['nombre']=datos_colaborador.usuario
            datos_nuevos['dpi']=int(float(datos_colaborador.dpi)) if datos_colaborador.dpi else ""
            datos_nuevos['edad']=datos_colaborador.edad
            for datos_enviar in datos_buscar_colaborador:
                datos_corregidos={}
                if datos_enviar.usuario is None:
                    datos_corregidos['fecha']=datos_enviar.fecha.date()
                    datos_corregidos['nombre']=datos_nuevos['nombre']
                    datos_corregidos['correo']=datos_enviar.email
                    datos_corregidos['dpi']=datos_nuevos['dpi']
                    datos_corregidos['edad']=datos_nuevos['edad']
                    if datos_enviar.pregunta4 =='Sí':
                        datos_corregidos['respuesta_si']=1
                        datos_corregidos['respuesta_no']=0
                    else:
                        datos_corregidos['respuesta_no']=1
                        datos_corregidos['respuesta_si']=0
                    datos_corregidos['pregunta1']=datos_enviar.pregunta1
                    datos_corregidos['pregunta2']=datos_enviar.pregunta2
                    datos_corregidos['pregunta3']=datos_enviar.pregunta3
                    datos_corregidos['pregunta4']=datos_enviar.pregunta4
                    datos_corregidos['pregunta5']=datos_enviar.pregunta5
                    datos_corregidos['pregunta6']=datos_enviar.pregunta6
                    datos_corregidos['pregunta7']=datos_enviar.pregunta7
                    datos_corregidos['pregunta8']=datos_enviar.pregunta8
                    datos_corregidos['pregunta9']=datos_enviar.pregunta9
                    datos_corregidos['pregunta10']=datos_enviar.pregunta10
                    datos_corregidos['pregunta11']=datos_enviar.pregunta11
                    datos_corregidos['pregunta12']=datos_enviar.pregunta12
                    datos_corregidos['pregunta13']=datos_enviar.pregunta13
                    datos_corregidos['pregunta14']=datos_enviar.pregunta14
                    datos_corregidos['pregunta15']=datos_enviar.pregunta15
                    datos_corregidos['pregunta16']=datos_enviar.pregunta16
                    datos_corregidos['pregunta17']=datos_enviar.pregunta17
                    datos_enviar_json.append(datos_corregidos)
                else:
                    datos_corregidos['fecha']=datos_enviar.fecha.date()
                    datos_corregidos['nombre']=datos_enviar.usuario
                    datos_corregidos['dpi']=int(float(datos_colaborador.dpi))
                    datos_corregidos['edad']=datos_enviar.edad
                    datos_corregidos['correo']=datos_enviar.email
                    if datos_enviar.pregunta4 =='Sí':
                        datos_corregidos['respuesta_si']=1
                        datos_corregidos['respuesta_no']=0
                    else:
                        datos_corregidos['respuesta_no']=1
                        datos_corregidos['respuesta_si']=0
                    datos_corregidos['pregunta1']=datos_enviar.pregunta1
                    datos_corregidos['pregunta2']=datos_enviar.pregunta2
                    datos_corregidos['pregunta3']=datos_enviar.pregunta3
                    datos_corregidos['pregunta4']=datos_enviar.pregunta4
                    datos_corregidos['pregunta5']=datos_enviar.pregunta5
                    datos_corregidos['pregunta6']=datos_enviar.pregunta6
                    datos_corregidos['pregunta7']=datos_enviar.pregunta7
                    datos_corregidos['pregunta8']=datos_enviar.pregunta8
                    datos_corregidos['pregunta9']=datos_enviar.pregunta9
                    datos_corregidos['pregunta10']=datos_enviar.pregunta10
                    datos_corregidos['pregunta11']=datos_enviar.pregunta11
                    datos_corregidos['pregunta12']=datos_enviar.pregunta12
                    datos_corregidos['pregunta13']=datos_enviar.pregunta13
                    datos_corregidos['pregunta14']=datos_enviar.pregunta14
                    datos_corregidos['pregunta15']=datos_enviar.pregunta15
                    datos_corregidos['pregunta16']=datos_enviar.pregunta16
                    datos_corregidos['pregunta17']=datos_enviar.pregunta17
                    datos_enviar_json.append(datos_corregidos)
            return Response(
                datos_enviar_json,
                status=status.HTTP_200_OK
            )

class GraficasBienestarJson(LoginRequiredMixin, views.APIView):
        """ Regreso los datos obtenidos del modelo de `Bienestar` para generar el informe
        """
        def get(self, request):
            contador_si_pregunta1=0
            contador_no_pregunta1=0
            contador_si_pregunta3=0
            contador_no_pregunta3=0
            contador_si_pregunta4 = 0
            contador_no_pregunta4 = 0
            contador_si_pregunta5 = 0
            contador_no_pregunta5 = 0
            contador_excelente_pregunta7 = 0
            contador_bueno_pregunta7 = 0
            contador_regular_pregunta7 = 0
            contador_malo_pregunta7 = 0
            contador_si_pregunta9 = 0
            contador_no_pregunta9 = 0
            contador_si_pregunta10 = 0
            contador_no_pregunta10 = 0
            contador_si_pregunta12 = 0
            contador_no_pregunta12 = 0
            contador_si_pregunta14 = 0
            contador_no_pregunta14 = 0
            contador_si_pregunta15 = 0
            contador_no_pregunta15 = 0
            contador_si_pregunta16 = 0
            contador_no_pregunta16 = 0
            datos_graficas={}
            datos_enviar=[]
            #pregunta1
            #¿ Padeces de alguna enfermedad o condición que te coloque en riesgo?
            pregunta_1=Colaborador.objects.values_list("pregunta1",flat=True)
            for respuesta_pregunta1 in pregunta_1:
                if respuesta_pregunta1 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta1 == "Sí":
                    #print("Encontramos un Si")
                    contador_si_pregunta1 = contador_si_pregunta1 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta1 = contador_no_pregunta1 + 1
            datos_graficas["contador_no_pregunta1"]=contador_no_pregunta1
            datos_graficas["contador_si_pregunta1"]=contador_si_pregunta1
            #pregunta3
            #¿Tienes familiares que vivan contigo en esta época?
            pregunta_3=Colaborador.objects.values_list("pregunta3",flat=True)
            for respuesta_pregunta3 in pregunta_3:
                if respuesta_pregunta3 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta3 == "Sí":
                    #print("Encontramos un Si")
                    contador_si_pregunta3 = contador_si_pregunta3 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta3 = contador_no_pregunta3 + 1
            datos_graficas["contador_no_pregunta3"]=contador_no_pregunta3
            datos_graficas["contador_si_pregunta3"]=contador_si_pregunta3
            #pregunta4
            #¿ Has presentado fiebre, tos, dolor de garganta, síntomas gastrointestinales (diarrea y/o vómito)  o dificultad para respirar en la últimas 24 horas?
            pregunta_4=Colaborador.objects.values_list("pregunta4",flat=True)
            for respuesta_pregunta4 in pregunta_4:
                if respuesta_pregunta4 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta4 == "Sí":
                    #print("Encontramos un Si")
                    contador_si_pregunta4 = contador_si_pregunta4 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta4 = contador_no_pregunta4 + 1
            datos_graficas["contador_no_pregunta4"]=contador_no_pregunta4
            datos_graficas["contador_si_pregunta4"]=contador_si_pregunta4
            #pregunta5
             #¿Ha cambiado tu situación de salud desde la última vez que respondiste este formulario (solo responder del segundo en adelante)?
            pregunta_5=Colaborador.objects.values_list("pregunta5",flat=True)
            for respuesta_pregunta5 in pregunta_5:
                 if respuesta_pregunta5 is None:
                     print("aca hay uno vacio")
                 elif respuesta_pregunta5 == "Sí":
                     #print("Encontramos un Si")
                     contador_si_pregunta5 = contador_si_pregunta5 + 1
                 else:
                     #print("Esto es un NO")
                     contador_no_pregunta5 = contador_no_pregunta5 + 1
            datos_graficas["contador_no_pregunta5"]=contador_no_pregunta5
            datos_graficas["contador_si_pregunta5"]=contador_si_pregunta5
            #pregunta7
            #¿Cómo calificas tu estado emocional en este momento?
            pregunta_7=Colaborador.objects.values_list("pregunta7",flat=True)
            for respuesta_pregunta7 in pregunta_7:
                if respuesta_pregunta7 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta7 == "Excelente":
                    contador_excelente_pregunta7 = contador_excelente_pregunta7 + 1
                elif respuesta_pregunta7 == "Bueno":
                    contador_bueno_pregunta7 = contador_bueno_pregunta7 + 1
                elif respuesta_pregunta7 == "Regular":
                    contador_regular_pregunta7 = contador_regular_pregunta7 + 1
                else:
                    contador_malo_pregunta7 = contador_malo_pregunta7 + 1
            datos_graficas["contador_excelente_pregunta7"]=contador_excelente_pregunta7
            datos_graficas["contador_bueno_pregunta7"]=contador_bueno_pregunta7
            datos_graficas["contador_regular_pregunta7"]=contador_regular_pregunta7
            datos_graficas["contador_malo_pregunta7"]=contador_malo_pregunta7
            #pregunta9
            #¿Has registrado un cambio en cuanto al número de personas que viven contigo?
            pregunta_9=Colaborador.objects.values_list("pregunta9",flat=True)
            for respuesta_pregunta9 in pregunta_9:
                if respuesta_pregunta9 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta9 == "Sí" or respuesta_pregunta9=="Si":
                    #print("Encontramos un Si")
                    contador_si_pregunta9 = contador_si_pregunta9 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta9 = contador_no_pregunta9 + 1
            datos_graficas["contador_no_pregunta9"]=contador_no_pregunta9
            datos_graficas["contador_si_pregunta9"]=contador_si_pregunta9
            #pregunta10
            #¿Tienes algún familiar o personas en casa que en las últimas 24 horas haya presentado alguno de los síntomas anteriores?
            pregunta_10=Colaborador.objects.values_list("pregunta10",flat=True)
            for respuesta_pregunta10 in pregunta_10:
                if respuesta_pregunta10 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta10 == "Sí":
                    #print("Encontramos un Si")
                    contador_si_pregunta10 = contador_si_pregunta10 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta10 = contador_no_pregunta10 + 1
            datos_graficas["contador_no_pregunta10"]=contador_no_pregunta10
            datos_graficas["contador_si_pregunta10"]=contador_si_pregunta10
            #pregunta12
            #¿Tuviste contacto con algún caso confirmado o sospechoso de COVID-19?
            pregunta_12=Colaborador.objects.values_list("pregunta12",flat=True)
            for respuesta_pregunta12 in pregunta_12:
                if respuesta_pregunta12 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta12 == "Sí":
                    #print("Encontramos un Si")
                    contador_si_pregunta12 = contador_si_pregunta12 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta12 = contador_no_pregunta12 + 1
            datos_graficas["contador_no_pregunta12"]=contador_no_pregunta12
            datos_graficas["contador_si_pregunta12"]=contador_si_pregunta12
            #pregunta14
            #¿Tiene tu colonia, municipio o comunidad donde vives cordón sanitario?
            pregunta_14=Colaborador.objects.values_list("pregunta14",flat=True)
            for respuesta_pregunta14 in pregunta_14:
                if respuesta_pregunta14 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta14 == "Sí":
                    #print("Encontramos un Si")
                    contador_si_pregunta14 = contador_si_pregunta14 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta14 = contador_no_pregunta14 + 1
            datos_graficas["contador_no_pregunta14"]=contador_no_pregunta14
            datos_graficas["contador_si_pregunta14"]=contador_si_pregunta14
            #pregunta15
            #¿ Tienes familiares o personas que vivan contigo en una casa que tenga labores de alta exposición de contacto? (repartidores de alimentos, medicina, o recepcionista?
            pregunta_15=Colaborador.objects.values_list("pregunta15",flat=True)
            for respuesta_pregunta15 in pregunta_15:
                if respuesta_pregunta15 is None:
                    print("aca hay uno vacio")
                elif respuesta_pregunta15 == "Sí":
                    #print("Encontramos un Si")
                    contador_si_pregunta15 = contador_si_pregunta15 + 1
                else:
                    #print("Esto es un NO")
                    contador_no_pregunta15 = contador_no_pregunta15 + 1
            datos_graficas["contador_no_pregunta15"]=contador_no_pregunta15
            datos_graficas["contador_si_pregunta15"]=contador_si_pregunta15
            #pregunta16
            #¿ Estás laborando actualmente en las instalaciones de la fundación?
            pregunta_16=Colaborador.objects.values_list("pregunta16",flat=True)
            for respuesta_pregunta16 in pregunta_16:
               if respuesta_pregunta16 is None:
                   print("aca hay uno vacio")
               elif respuesta_pregunta16 == "Sí":
                   #print("Encontramos un Si")
                   contador_si_pregunta16 = contador_si_pregunta16 + 1
               else:
                   #print("Esto es un NO")
                   contador_no_pregunta16 = contador_no_pregunta16 + 1
            datos_graficas["contador_no_pregunta16"]=contador_no_pregunta16
            datos_graficas["contador_si_pregunta16"]=contador_si_pregunta16
            datos_enviar.append(datos_graficas)
            return Response(
                datos_enviar,
                status=status.HTTP_200_OK
              )
class InformeLineaTiempoTodosBienestarJson(LoginRequiredMixin, views.APIView):
        """ Regreso los datos obtenidos del modelo de `Bienestar` para generar el informe
        """
        def get(self, request):
            datos_enviar_json=[]
            fecha_sin_repetir=[]
            fecha_sin_repetir_buscar=[]
            datos_colaborador=Colaborador.objects.all().order_by('-fecha').values('fecha').distinct()
            file_list =os.listdir(settings.MEDIA_ROOT_EXCEL)
            number_file=len(file_list)
            print(number_file)
            for nueva_fecha in datos_colaborador:
                x=dateutil.parser.parse(str(nueva_fecha['fecha'])).date()
                fecha_sin_repetir.append(x)
            fecha_sin_repetir_buscar=sorted(list(set(fecha_sin_repetir)))
            for datos_nuevos in fecha_sin_repetir_buscar:
                datos_nuevos_enviar={}
                no=Colaborador.objects.filter(pregunta4='No',fecha__startswith=datos_nuevos).count()
                si=Colaborador.objects.filter(pregunta4='Sí',fecha__startswith=datos_nuevos).count()
                datos_nuevos_enviar['fecha']=str(datos_nuevos)
                datos_nuevos_enviar['si']=si
                datos_nuevos_enviar['no']=no
                datos_enviar_json.append(datos_nuevos_enviar)
            return Response(
            datos_enviar_json,
                status=status.HTTP_200_OK
            )
class InformeBienestarPieJson(LoginRequiredMixin, views.APIView):
        """ Regreso los datos obtenidos del modelo de `Bienestar` para generar el informe
        """

        def get(self, request):
            def miFuncion(datos):
                datos_enviar_json=[]
                for datos_enviar in datos:
                    datos_corregidos={}
                    if datos_enviar.usuario is None:
                        datos_colaborador=Colaborador.objects.filter(email=datos_enviar.email).first()
                        datos_corregidos['fecha']=datos_enviar.fecha.date()
                        datos_corregidos['nombre']=datos_colaborador.usuario
                        datos_corregidos['correo']=datos_enviar.email
                        datos_corregidos['dpi']=datos_colaborador.dpi
                        datos_corregidos['edad']=datos_colaborador.edad
                        datos_corregidos['pregunta1']=datos_enviar.pregunta1
                        datos_corregidos['pregunta2']=datos_enviar.pregunta2
                        datos_corregidos['pregunta3']=datos_enviar.pregunta3
                        datos_corregidos['pregunta4']=datos_enviar.pregunta4
                        datos_corregidos['pregunta5']=datos_enviar.pregunta5
                        datos_corregidos['pregunta6']=datos_enviar.pregunta6
                        datos_corregidos['pregunta7']=datos_enviar.pregunta7
                        datos_corregidos['pregunta8']=datos_enviar.pregunta8
                        datos_corregidos['pregunta9']=datos_enviar.pregunta9
                        datos_corregidos['pregunta10']=datos_enviar.pregunta10
                        datos_corregidos['pregunta11']=datos_enviar.pregunta11
                        datos_corregidos['pregunta12']=datos_enviar.pregunta12
                        datos_corregidos['pregunta13']=datos_enviar.pregunta13
                        datos_corregidos['pregunta14']=datos_enviar.pregunta14
                        datos_corregidos['pregunta15']=datos_enviar.pregunta15
                        datos_corregidos['pregunta16']=datos_enviar.pregunta16
                        datos_corregidos['pregunta17']=datos_enviar.pregunta17
                        datos_enviar_json.append(datos_corregidos)
                    else:
                        datos_corregidos['fecha']=datos_enviar.fecha.date()
                        datos_corregidos['nombre']=datos_enviar.usuario
                        datos_corregidos['dpi']=int(float(datos_enviar.dpi))
                        datos_corregidos['edad']=datos_enviar.edad
                        datos_corregidos['correo']=datos_enviar.email
                        datos_corregidos['pregunta1']=datos_enviar.pregunta1
                        datos_corregidos['pregunta2']=datos_enviar.pregunta2
                        datos_corregidos['pregunta3']=datos_enviar.pregunta3
                        datos_corregidos['pregunta4']=datos_enviar.pregunta4
                        datos_corregidos['pregunta5']=datos_enviar.pregunta5
                        datos_corregidos['pregunta6']=datos_enviar.pregunta6
                        datos_corregidos['pregunta7']=datos_enviar.pregunta7
                        datos_corregidos['pregunta8']=datos_enviar.pregunta8
                        datos_corregidos['pregunta9']=datos_enviar.pregunta9
                        datos_corregidos['pregunta10']=datos_enviar.pregunta10
                        datos_corregidos['pregunta11']=datos_enviar.pregunta11
                        datos_corregidos['pregunta12']=datos_enviar.pregunta12
                        datos_corregidos['pregunta13']=datos_enviar.pregunta13
                        datos_corregidos['pregunta14']=datos_enviar.pregunta14
                        datos_corregidos['pregunta15']=datos_enviar.pregunta15
                        datos_corregidos['pregunta16']=datos_enviar.pregunta16
                        datos_corregidos['pregunta17']=datos_enviar.pregunta17
                        datos_enviar_json.append(datos_corregidos)
                return datos_enviar_json
            if(self.request.GET['pregunta']=="pregunta1"):
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta1='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta1='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta3"):
                print("Ingreso a la pregunta 3")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta3='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta3='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta4"):
                print("Ingreso a la pregunta 4")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta4='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta4='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta5"):
                print("Ingreso a la pregunta 5")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta5='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta5='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta7"):
                print("Ingreso a la pregunta 7")
                if(self.request.GET['respuesta']=="Excelente"):
                    datos=Colaborador.objects.filter(pregunta7='Excelente')
                    informe=miFuncion(datos)
                elif(self.request.GET['respuesta']=="Bueno"):
                    datos=Colaborador.objects.filter(pregunta7='Bueno')
                    informe=miFuncion(datos)
                elif(self.request.GET['respuesta']=="Regular"):
                    datos=Colaborador.objects.filter(pregunta7='Regular')
                    informe=miFuncion(datos)
                elif(self.request.GET['respuesta']=="Malo"):
                    datos=Colaborador.objects.filter(pregunta7='Malo')
                    informe=miFuncion(datos)


            elif(self.request.GET['pregunta']=="pregunta9"):
                print("Ingreso a la pregunta 9")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta9='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta9='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta10"):
                print("Ingreso a la pregunta 10")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta10='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta10='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta12"):
                print("Ingreso a la pregunta 12")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta12='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta12='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta14"):
                print("Ingreso a la pregunta 14")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta14='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta14='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta15"):
                print("Ingreso a la pregunta 15")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta15='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta15='Sí')
                    informe=miFuncion(datos)
            elif(self.request.GET['pregunta']=="pregunta16"):
                print("Ingreso a la pregunta 16")
                if(self.request.GET['respuesta']=="No"):
                    datos=Colaborador.objects.filter(pregunta16='No')
                    informe=miFuncion(datos)
                else:
                    datos=Colaborador.objects.filter(pregunta16='Sí')
                    informe=miFuncion(datos)
            return Response(
            informe,
                status=status.HTTP_200_OK
            )
class InformeBienestarTodosJson(LoginRequiredMixin, views.APIView):
        """ Regreso los datos obtenidos del modelo de `Bienestar` para generar el informe
        """
        def get(self, request):
            datos_buscar_colaborador =Colaborador.objects.all()
            datos_enviar_json=[]
            for datos_enviar in datos_buscar_colaborador:
                datos_corregidos={}
                datos_corregidos['fecha']=datos_enviar.fecha.date()
                datos_corregidos['nombre']=datos_enviar.usuario
                datos_corregidos['dpi']=datos_enviar.dpi
                datos_corregidos['edad']=datos_enviar.edad
                datos_corregidos['correo']=datos_enviar.email
                datos_corregidos['pregunta1']=datos_enviar.pregunta1
                datos_corregidos['pregunta2']=datos_enviar.pregunta2
                datos_corregidos['pregunta3']=datos_enviar.pregunta3
                datos_corregidos['pregunta4']=datos_enviar.pregunta4
                datos_corregidos['pregunta5']=datos_enviar.pregunta5
                datos_corregidos['pregunta6']=datos_enviar.pregunta6
                datos_corregidos['pregunta7']=datos_enviar.pregunta7
                datos_corregidos['pregunta8']=datos_enviar.pregunta8
                datos_corregidos['pregunta9']=datos_enviar.pregunta9
                datos_corregidos['pregunta10']=datos_enviar.pregunta10
                datos_corregidos['pregunta11']=datos_enviar.pregunta11
                datos_corregidos['pregunta12']=datos_enviar.pregunta12
                datos_corregidos['pregunta13']=datos_enviar.pregunta13
                datos_corregidos['pregunta14']=datos_enviar.pregunta14
                datos_corregidos['pregunta15']=datos_enviar.pregunta15
                datos_corregidos['pregunta16']=datos_enviar.pregunta16
                datos_corregidos['pregunta17']=datos_enviar.pregunta17
                datos_enviar_json.append(datos_corregidos)
            return Response(
                datos_enviar_json,
                status=status.HTTP_200_OK
            )
