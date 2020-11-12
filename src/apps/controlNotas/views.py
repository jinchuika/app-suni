from django.shortcuts import render
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView, FormView
from django.urls import reverse_lazy
from openpyxl import load_workbook
from django.http import HttpResponse, HttpResponseRedirect
from rest_framework import views, status
from rest_framework.response import Response
from django.core.exceptions import ObjectDoesNotExist
from braces.views import (
    LoginRequiredMixin, GroupRequiredMixin)
import os
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import json
from apps.escuela.models import Escuela
from apps.controlNotas import models as control_m
from apps.controlNotas import forms as control_f
from django.db.models import Avg, Count, Min, Sum
# Create your views here.
class ControlExcelAddView( TemplateView):
    template_name = 'controlNotas/tablasNota.html'

class RegistrosExcelAddView(LoginRequiredMixin, TemplateView, GroupRequiredMixin):
    template_name = 'controlNotas/cargar_excel.html'
    def post(self,request):
        if request.method == 'POST' and request.FILES['myfile']:
            file_list =os.listdir(settings.MEDIA_ROOT_EXCEL_IMPACTO)
            number_file=len(file_list)
            myfile=request.FILES['myfile']
            new_name=str('Impacto') + str(number_file) +str(".")+str(myfile.name.split(".")[-1])
            fs=FileSystemStorage(location=settings.MEDIA_ROOT_EXCEL_IMPACTO)
            filename=fs.save(new_name, myfile)
            uploaded_file_url=fs.url(filename)
            return render(request, 'controlNotas/cargar_excel.html',{
                'uploaded_file_url':uploaded_file_url
        })
        return render(request,'ControlNotas/cargar_excel.html')

class RegistrosAddView(CreateView, GroupRequiredMixin, LoginRequiredMixin):
    group_required = [u"mye",]
    template_name = 'controlNotas/registros.html'
    model = control_m.Visita
    form_class = control_f.EvaluacionForm
    def get_context_data(self, **kwargs):
        contador_materias=0
        id_contador=0
        progreso=0
        notas_enviar=[]
        alumno_enviar=[]
        materias_enviar=[]
        data_enviar=[]
        context = super(RegistrosAddView, self).get_context_data(**kwargs)
        nombre=control_m.Visita.objects.all().last()
        nombre_escuela= str(nombre.escuela) + str("-")+str(nombre.semestre)
        context['escuela'] = nombre_escuela
        context['escuela_nombre'] = str(nombre.escuela)
        #materia =control_m.Notas.objects.filter(evaluacion__visita__escuela__id=nombre.escuela.id,evaluacion__visita__semestre=nombre.semestre).values('evaluacion__materia').distinct()
        materia =control_m.Notas.objects.filter(evaluacion__visita__escuela__id=nombre.escuela.id,evaluacion__visita__semestre=nombre.semestre).values('evaluacion__materia','evaluacion__grado__nombre_grado').distinct()
        for data in materia:
            datos_enviar={}
            materia =control_m.Materia.objects.filter(id=data["evaluacion__materia"]).values('nombre','icon','id','color')
            contador_materias = contador_materias + materia.count()
            id_contador = id_contador + 1
            datos_enviar["icono"]=materia[0]['icon']
            datos_enviar["materias"]=materia[0]['nombre']
            datos_enviar["id"]=materia[0]['id']
            datos_enviar["color"]=materia[0]['color']
            datos_enviar["grado"]=data["evaluacion__grado__nombre_grado"]
            notas = control_m.Notas.objects.filter(evaluacion__materia=data["evaluacion__materia"]).aggregate(total_util=Avg('nota'))
            participantes =control_m.Notas.objects.filter(evaluacion__materia=data["evaluacion__materia"]).values('alumno','nota','evaluacion__grado__nombre_grado',"evaluacion__grado")
            progreso= progreso + notas["total_util"]
            for value in participantes:
                datos_participante ={}
                datos_participante["alumno"]=value['alumno']
                datos_participante["nota"]=round(value['nota'],2)                
                alumno_enviar.append(datos_participante)
            datos_enviar["alumno"]=alumno_enviar
            alumno_enviar=[]
            notas_enviar=[]
            context['escuela_porcentaje'] = round((progreso / id_contador),2)
            datos_enviar["notas"]=round(notas["total_util"],2)
            datos_enviar["id_contador"]=id_contador
            data_enviar.append(datos_enviar)
        context["datos"]=data_enviar
        context["numero_materias"]=contador_materias
        return context

class ResultadoNotasJson(LoginRequiredMixin, views.APIView):
        """ Importar registros de excel con DjangoRestFramework para Bienestar
        """
        def get(self, request):
            datos_enviar=[]
            file_list =os.listdir(settings.MEDIA_ROOT_EXCEL_IMPACTO)
            number_file=len(file_list)
            name_file =file_list[number_file-1]
            ruta = str(settings.MEDIA_ROOT_EXCEL_IMPACTO)+str(name_file)
            listado_datos=[]
            wb_obj = load_workbook(filename = ruta)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            m_col= sheet_obj.max_column
            for i in range(1, m_row+1):
            #for i in range(2, m_row+1):
                nombre=sheet_obj.cell(row=i, column=1).value
                recolectar_datos = {}
                recolectar_datos["nombre"]=nombre
                datos_enviar.append(recolectar_datos)
            return Response(
                datos_enviar,
                status=status.HTTP_200_OK
            )
        def post(self, request):
            visita=control_m.Visita.objects.all().last()
            materia=request.data["materia"]
            grado=request.data["grado"]
            observacion=request.data["observacion"]
            nueva_evaluacion = control_m.Evaluacion(
                visita=visita,
                materia=control_m.Materia.objects.get(id=materia),
                grado=control_m.Grado.objects.get(id=grado),
                observacion=observacion
            )
            nueva_evaluacion.save()
            evaluacion=control_m.Evaluacion.objects.all().last()
            dato =  json.loads(request.data['datos'])
            for info in dato:
                nota = control_m.Notas(
                    evaluacion=evaluacion,
                    alumno=info["nombre"],
                    nota=info["nota"]
                )
                nota.save()
            return Response(
                "Notas Guardadas Exitosamente",
                status=status.HTTP_200_OK

            )
class VisitasAddView(LoginRequiredMixin, views.APIView):
        """ Importar registros de excel con DjangoRestFramework para Bienestar
        """
        def post(self, request):
            escuela=Escuela.objects.get(codigo=self.request.POST['escuela'])
            semestre=control_m.Semestre.objects.get(numero=self.request.POST['semestre'])
            numero_visitas = control_m.Visita.objects.all().count()
            if numero_visitas >=1:
                nueva_visita = control_m.Visita(
                    escuela=escuela,
                    semestre=semestre,
                    usuario=self.request.user,
                    numero_visita=numero_visitas+1
                )
            else:
                nueva_visita = control_m.Visita(
                    escuela=escuela,
                    semestre=semestre,
                    usuario=self.request.user,
                    numero_visita=1
                )
            nueva_visita.save()
            return Response(
                "Visita creada exitosamente",
                status=status.HTTP_200_OK

            )
        def get(self, request):
            datos_enviar=[]
            visitas=control_m.Visita.objects.filter(escuela__codigo=self.request.GET['escuela'])
            for visita in visitas:
                datos_linea_tiempo={}
                materias=control_m.Evaluacion.objects.filter(visita=visita).values("materia__nombre").distinct()
                notas =control_m.Notas.objects.filter(evaluacion__visita=visita).aggregate(promedio=Avg('nota'))
                datos_linea_tiempo["visita"]=str(visita.fecha.year)+str("-")+str(visita.semestre)
                datos_linea_tiempo["promedio"]=round(notas["promedio"],0)
                datos_enviar.append(datos_linea_tiempo)
            return Response(
                datos_enviar,
                status=status.HTTP_200_OK

            )
